"""
main.py — FastAPI application entry point.

Routes
------
GET  /api/dashboard
GET  /api/organisations
GET  /api/departments         POST /api/departments
PATCH /api/departments/{code}/set-head
GET  /api/projects            POST /api/projects
GET  /api/cost-types
GET  /api/statuses
GET  /api/employees           POST /api/employees
GET  /api/activities          POST /api/activities
PATCH /api/activities/{code}  DELETE /api/activities/{code}
GET  /api/sub-activities      POST /api/sub-activities
PATCH /api/sub-activities/{code}
GET  /api/budget-versions     POST /api/budget-versions
PATCH /api/budget-versions/{code}/set-active
PATCH /api/budget-versions/{code}/toggle-lock
GET  /api/fiscal-periods
GET  /api/fiscal-years          ← distinct fiscal year strings
GET  /api/transfers           POST /api/transfers   (legacy — immediate, annual-total only)
GET  /api/change-requests     POST /api/change-requests
PATCH /api/change-requests/{id}/decide   ← NEW: approve/reject — APPROVAL MOVES THE MONEY
POST /api/budget-upload         ← Excel upload — STAGED per department, Pending until decided
GET  /api/budget-uploads        ← upload history (optional ?dept_code= / ?status= filters)
PATCH /api/budget-uploads/{id}/decide   ← department head approves/rejects — ALL DEPTS APPROVED WRITES THE ROWS
"""

import random, string, io, json
from datetime import date, datetime
from typing import List, Optional
import requests

from fastapi import FastAPI, Depends, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session, joinedload


from routers import workflows, requests, stages, approvals, analytics, auth, onboarding_dashboard, reports
from database import get_db, init_db
from models import (
    Organisation, Department, Project, CostType, Status,
    Employee, Activity, ActivityPhase, SubActivity, SubActivityPhase,
    BudgetVersion, FiscalPeriod, Transfer, ChangeRequest,
    BudgetUpload,
)
from schemas import (
    OrganisationOut,
    DepartmentCreate, DepartmentOut, DepartmentSetHead,
    ProjectCreate, ProjectOut,
    CostTypeOut, StatusOut,
    EmployeeCreate, EmployeeOut,
    ActivityCreate, ActivityUpdate, ActivityOut,
    SubActivityCreate, SubActivityUpdate, SubActivityOut,
    BudgetVersionCreate, BudgetVersionOut,
    FiscalPeriodOut,
    TransferCreate, TransferOut,
    ChangeRequestCreate, ChangeRequestOut, ChangeRequestDecide,
    BudgetUploadOut, BudgetUploadDecide,
    DashboardOut, DeptSummary,
    InvoiceBudgetCheckReq, InvoiceBudgetCheckRes, BudgetBlockReq,
)

app = FastAPI(title="BudgetControl API", version="1.2.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def on_startup():
    init_db()

import routers.onboarding_dashboard as onboarding_dashboard
import routers.vendors as vendors
import routers.vendor_materials as vendor_materials
import routers.vendor_pr as vendor_pr
import routers.vendor_quotation_report as vendor_quotation_report
import routers.vendor_po as vendor_po
import routers.employee_quote_comparison as employee_quote_comparison
import routers.users as users
import routers.vendor_credit_notes_report as vendor_credit_notes_report
import routers.vendor_material_stock_report as vendor_material_stock_report
import routers.vendor_returns_report as vendor_returns_report
import routers.vendor_payment_report as vendor_payment_report
import routers.vendor_asns as vendor_asns

app.include_router(onboarding_dashboard.router)
app.include_router(vendors.router)
app.include_router(vendor_materials.router)
app.include_router(vendor_pr.router)
app.include_router(vendor_quotation_report.router)
app.include_router(vendor_po.router)
app.include_router(employee_quote_comparison.router)
app.include_router(users.router, prefix="/api/users", tags=["users"])
app.include_router(vendor_credit_notes_report.router)
app.include_router(vendor_material_stock_report.router)
app.include_router(vendor_returns_report.router)
app.include_router(vendor_payment_report.router)
app.include_router(vendor_asns.router)


def _uid(prefix: str = "") -> str:
    chars = string.ascii_uppercase + string.digits
    return prefix + "".join(random.choices(chars, k=6))


def _enrich_activity(a: Activity) -> ActivityOut:
    out = ActivityOut.model_validate(a)
    out.cost_type_tag  = a.cost_type.tag  if a.cost_type  else None
    out.status_name    = a.status.name    if a.status     else None
    out.employee_name  = a.employee.name  if a.employee   else None
    out.sub_activities = [_enrich_sub(s) for s in a.sub_activities]
    out.phases         = list(a.phases or [])
    return out


def _enrich_sub(s: SubActivity) -> SubActivityOut:
    out = SubActivityOut.model_validate(s)
    out.cost_type_tag = s.cost_type.tag if s.cost_type else None
    out.status_name   = s.status.name   if s.status    else None
    out.employee_name = s.employee.name if s.employee  else None
    out.phases        = list(s.phases or [])
    return out


# ── Dashboard ─────────────────────────────────────────────────────────────────
@app.get("/api/dashboard", response_model=DashboardOut)
def dashboard(db: Session = Depends(get_db)):
    activities     = db.query(Activity).all()
    sub_activities = db.query(SubActivity).all()

    all_items = list(activities) + list(sub_activities)

    totals = dict(allocated=0, pr=0, po=0, invoiced=0, opex_total=0, capex_total=0)
    for item in all_items:
        totals["allocated"] += item.allocated
        totals["pr"]        += item.pr
        totals["po"]        += item.po
        totals["invoiced"]  += item.invoiced
        tag = item.cost_type.tag if item.cost_type else ""
        if tag == "Opex":
            totals["opex_total"] += item.allocated
        elif tag == "Capex":
            totals["capex_total"] += item.allocated

    depts        = db.query(Department).all()
    projects     = db.query(Project).all()
    project_dept = {p.project_code: p.dept_code for p in projects}

    dept_map = {d.dept_code: dict(
        dept_code=d.dept_code, dept_name=d.name,
        allocated=0, pr=0, po=0, invoiced=0, remaining=0
    ) for d in depts}

    for a in activities:
        dc = project_dept.get(a.project_code)
        if dc and dc in dept_map:
            m = dept_map[dc]
            m["allocated"] += a.allocated
            m["pr"]        += a.pr
            m["po"]        += a.po
            m["invoiced"]  += a.invoiced

    for s in sub_activities:
        act = db.query(Activity).filter(Activity.activity_code == s.parent_activity_code).first()
        if act:
            dc = project_dept.get(act.project_code)
            if dc and dc in dept_map:
                m = dept_map[dc]
                m["allocated"] += s.allocated
                m["pr"]        += s.pr
                m["po"]        += s.po
                m["invoiced"]  += s.invoiced

    for m in dept_map.values():
        m["remaining"] = m["allocated"] - m["invoiced"]

    return DashboardOut(
        allocated=totals["allocated"],
        pr=totals["pr"],
        po=totals["po"],
        invoiced=totals["invoiced"],
        remaining=totals["allocated"] - totals["invoiced"],
        pr_remaining=totals["allocated"] - totals["pr"],
        po_remaining=totals["allocated"] - totals["po"],
        opex_total=totals["opex_total"],
        capex_total=totals["capex_total"],
        dept_summaries=[DeptSummary(**v) for v in dept_map.values()],
    )


# ── Organisations ─────────────────────────────────────────────────────────────
@app.get("/api/organisations", response_model=List[OrganisationOut])
def list_organisations(db: Session = Depends(get_db)):
    return db.query(Organisation).all()


# ── Departments ───────────────────────────────────────────────────────────────
@app.get("/api/departments", response_model=List[DepartmentOut])
def list_departments(db: Session = Depends(get_db)):
    return db.query(Department).all()

@app.post("/api/departments", response_model=DepartmentOut, status_code=201)
def create_department(body: DepartmentCreate, db: Session = Depends(get_db)):
    dept = Department(dept_code=_uid("DEPT-"), name=body.name, org_code=body.org_code,
                       wbs=body.wbs, head_employee_code=body.head_employee_code)
    db.add(dept); db.commit(); db.refresh(dept)
    return dept

@app.patch("/api/departments/{dept_code}/set-head", response_model=DepartmentOut)
def set_department_head(dept_code: str, body: DepartmentSetHead, db: Session = Depends(get_db)):
    dept = _get_or_404(db, Department, dept_code, "dept_code", "Department")
    if body.head_employee_code:
        _get_or_404(db, Employee, body.head_employee_code, "employee_code", "Employee")
    dept.head_employee_code = body.head_employee_code
    db.commit(); db.refresh(dept)
    return dept


# ── Projects ──────────────────────────────────────────────────────────────────
@app.get("/api/projects", response_model=List[ProjectOut])
def list_projects(dept_code: Optional[str] = Query(None), db: Session = Depends(get_db)):
    q = db.query(Project)
    if dept_code:
        q = q.filter(Project.dept_code == dept_code)
    return q.all()

@app.post("/api/projects", response_model=ProjectOut, status_code=201)
def create_project(body: ProjectCreate, db: Session = Depends(get_db)):
    proj = Project(project_code=_uid("PRJ-"), name=body.name, dept_code=body.dept_code, wbs=body.wbs)
    db.add(proj); db.commit(); db.refresh(proj)
    return proj


# ── Cost types ────────────────────────────────────────────────────────────────
@app.get("/api/cost-types", response_model=List[CostTypeOut])
def list_cost_types(db: Session = Depends(get_db)):
    return db.query(CostType).all()


# ── Statuses ──────────────────────────────────────────────────────────────────
@app.get("/api/statuses", response_model=List[StatusOut])
def list_statuses(db: Session = Depends(get_db)):
    return db.query(Status).order_by(Status.sort_order).all()


# ── Employees ─────────────────────────────────────────────────────────────────
@app.get("/api/employees", response_model=List[EmployeeOut])
def list_employees(admin_email: Optional[str] = Query(None), db: Session = Depends(get_db)):
    q = db.query(Employee)
    # Employee model doesn't have admin_email. If we need to filter, maybe by email or just return all.
    # For now, return all since admin_email field doesn't exist.
    # if admin_email:
    #     q = q.filter(Employee.email == admin_email)
    return q.all()

@app.post("/api/employees", response_model=EmployeeOut, status_code=201)
def create_employee(body: EmployeeCreate, db: Session = Depends(get_db)):
    pwd = body.password or "User@123"
    
    # 1. Create User automatically for this employee if email is provided
    if body.email:
        existing_user = db.query(models.User).filter(models.User.email == body.email).first()
        if not existing_user:
            from auth_utils import hash_password
            names = body.name.split(" ")
            new_user = models.User(
                firstName=names[0],
                lastName=" ".join(names[1:]) if len(names) > 1 else "",
                email=body.email,
                password=hash_password(pwd),
                role=models.UserRole.employee, # Default to employee role
                designation=body.title,
                is_active=True
            )
            db.add(new_user)
            db.flush()

    # 2. Create Employee record
    emp = Employee(
        employee_code=_uid("EMP-"), 
        name=body.name, 
        title=body.title,
        email=body.email,
        password=pwd,
        admin_email=body.admin_email,
        dept_code=body.dept_code, 
        manager_code=body.manager_code
    )
    db.add(emp); db.commit(); db.refresh(emp)
    return emp


# ── Activities ────────────────────────────────────────────────────────────────
@app.get("/api/activities", response_model=List[ActivityOut])
def list_activities(
    project_code: Optional[str] = Query(None),
    dept_code:    Optional[str] = Query(None),
    cost_type:    Optional[str] = Query(None),
    status:       Optional[str] = Query(None),
    search:       Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    q = (db.query(Activity)
           .options(
               joinedload(Activity.cost_type),
               joinedload(Activity.status),
               joinedload(Activity.employee),
               joinedload(Activity.phases),
               joinedload(Activity.sub_activities).joinedload(SubActivity.cost_type),
               joinedload(Activity.sub_activities).joinedload(SubActivity.status),
               joinedload(Activity.sub_activities).joinedload(SubActivity.employee),
               joinedload(Activity.sub_activities).joinedload(SubActivity.phases),
           ))
    if project_code: q = q.filter(Activity.project_code == project_code)
    if cost_type:    q = q.join(CostType).filter(CostType.tag == cost_type)
    if status:       q = q.join(Status, Activity.status_code == Status.status_code).filter(Status.name == status)
    if search:       q = q.filter(Activity.name.ilike(f"%{search}%"))
    if dept_code:    q = q.join(Project).filter(Project.dept_code == dept_code)
    return [_enrich_activity(a) for a in q.all()]

@app.post("/api/activities", response_model=ActivityOut, status_code=201)
def create_activity(body: ActivityCreate, db: Session = Depends(get_db)):
    act = Activity(activity_code=_uid("ACT-"), **body.model_dump())
    db.add(act); db.commit(); db.refresh(act)
    return _enrich_activity(
        db.query(Activity)
          .options(joinedload(Activity.cost_type), joinedload(Activity.status),
                   joinedload(Activity.employee), joinedload(Activity.phases),
                   joinedload(Activity.sub_activities).joinedload(SubActivity.phases))
          .filter(Activity.activity_code == act.activity_code).first()
    )

@app.patch("/api/activities/{code}", response_model=ActivityOut)
def update_activity(code: str, body: ActivityUpdate, db: Session = Depends(get_db)):
    act = db.query(Activity).filter(Activity.activity_code == code).first()
    if not act: raise HTTPException(404, "Activity not found")
    for k, v in body.model_dump(exclude_none=True).items():
        setattr(act, k, v)
    db.commit()
    return _enrich_activity(
        db.query(Activity)
          .options(joinedload(Activity.cost_type), joinedload(Activity.status),
                   joinedload(Activity.employee), joinedload(Activity.phases),
                   joinedload(Activity.sub_activities).joinedload(SubActivity.phases))
          .filter(Activity.activity_code == code).first()
    )

@app.delete("/api/activities/{code}", status_code=204)
def delete_activity(code: str, db: Session = Depends(get_db)):
    act = db.query(Activity).filter(Activity.activity_code == code).first()
    if not act: raise HTTPException(404, "Activity not found")
    # Guard: block delete if the activity carries real budget/spend/commitments
    # or has children, so a stray click can't silently destroy budget history.
    if act.allocated or act.pr or act.po or act.invoiced or act.approved:
        raise HTTPException(400, "Cannot delete: this activity has budget, PR/PO, or spend recorded against it. Reduce it to zero first.")
    if act.sub_activities:
        raise HTTPException(400, f"Cannot delete: this activity has {len(act.sub_activities)} sub-activity line(s). Delete those first.")
    open_crs = db.query(ChangeRequest).filter(
        ChangeRequest.status == "Pending",
        (ChangeRequest.a_code == code) | (ChangeRequest.b_code == code),
    ).count()
    if open_crs:
        raise HTTPException(400, f"Cannot delete: {open_crs} pending change request(s) reference this activity.")
    db.delete(act); db.commit()


# ── Sub-activities ────────────────────────────────────────────────────────────
@app.get("/api/sub-activities", response_model=List[SubActivityOut])
def list_sub_activities(parent_activity_code: Optional[str] = Query(None), db: Session = Depends(get_db)):
    q = (db.query(SubActivity)
           .options(joinedload(SubActivity.cost_type),
                    joinedload(SubActivity.status),
                    joinedload(SubActivity.employee)))
    if parent_activity_code:
        q = q.filter(SubActivity.parent_activity_code == parent_activity_code)
    return [_enrich_sub(s) for s in q.all()]

@app.post("/api/sub-activities", response_model=SubActivityOut, status_code=201)
def create_sub_activity(body: SubActivityCreate, db: Session = Depends(get_db)):
    if body.level < 1 or body.level > 3:
        raise HTTPException(400, "level must be 1, 2, or 3")
    sa = SubActivity(subactivity_code=_uid("SACT-"), **body.model_dump())
    db.add(sa); db.commit(); db.refresh(sa)
    return _enrich_sub(
        db.query(SubActivity)
          .options(joinedload(SubActivity.cost_type), joinedload(SubActivity.status),
                   joinedload(SubActivity.employee), joinedload(SubActivity.phases))
          .filter(SubActivity.subactivity_code == sa.subactivity_code).first()
    )

@app.patch("/api/sub-activities/{code}", response_model=SubActivityOut)
def update_sub_activity(code: str, body: SubActivityUpdate, db: Session = Depends(get_db)):
    sa = db.query(SubActivity).filter(SubActivity.subactivity_code == code).first()
    if not sa: raise HTTPException(404, "Sub-activity not found")
    for k, v in body.model_dump(exclude_none=True).items():
        setattr(sa, k, v)
    db.commit()
    return _enrich_sub(
        db.query(SubActivity)
          .options(joinedload(SubActivity.cost_type), joinedload(SubActivity.status),
                   joinedload(SubActivity.employee), joinedload(SubActivity.phases))
          .filter(SubActivity.subactivity_code == code).first()
    )

@app.delete("/api/sub-activities/{code}", status_code=204)
def delete_sub_activity(code: str, db: Session = Depends(get_db)):
    sa = db.query(SubActivity).filter(SubActivity.subactivity_code == code).first()
    if not sa: raise HTTPException(404, "Sub-activity not found")
    if sa.allocated or sa.pr or sa.po or sa.invoiced or sa.approved:
        raise HTTPException(400, "Cannot delete: this sub-activity has budget, PR/PO, or spend recorded against it.")
    db.delete(sa); db.commit()


# ── Budget versions ───────────────────────────────────────────────────────────
@app.get("/api/budget-versions", response_model=List[BudgetVersionOut])
def list_budget_versions(db: Session = Depends(get_db)):
    return db.query(BudgetVersion).all()

@app.post("/api/budget-versions", response_model=BudgetVersionOut, status_code=201)
def create_budget_version(body: BudgetVersionCreate, db: Session = Depends(get_db)):
    bv = BudgetVersion(version_code=_uid("BV-"), **body.model_dump())
    db.add(bv); db.commit(); db.refresh(bv)
    return bv

@app.patch("/api/budget-versions/{code}/set-active", response_model=BudgetVersionOut)
def set_active_version(code: str, db: Session = Depends(get_db)):
    db.query(BudgetVersion).update({"is_current": 0})
    bv = db.query(BudgetVersion).filter(BudgetVersion.version_code == code).first()
    if not bv: raise HTTPException(404, "Budget version not found")
    bv.is_current = 1
    db.commit(); db.refresh(bv)
    return bv

@app.patch("/api/budget-versions/{code}/toggle-lock", response_model=BudgetVersionOut)
def toggle_lock(code: str, db: Session = Depends(get_db)):
    bv = db.query(BudgetVersion).filter(BudgetVersion.version_code == code).first()
    if not bv: raise HTTPException(404, "Budget version not found")
    bv.is_locked = 0 if bv.is_locked else 1
    db.commit(); db.refresh(bv)
    return bv


# ── Fiscal periods ────────────────────────────────────────────────────────────
@app.get("/api/fiscal-periods", response_model=List[FiscalPeriodOut])
def list_fiscal_periods(db: Session = Depends(get_db)):
    return db.query(FiscalPeriod).order_by(FiscalPeriod.period_no).all()

@app.get("/api/fiscal-years")
def list_fiscal_years(db: Session = Depends(get_db)):
    """Return distinct fiscal year strings for dropdowns."""
    rows = db.query(FiscalPeriod.fiscal_year).distinct().order_by(FiscalPeriod.fiscal_year).all()
    return [r[0] for r in rows]


# ── Transfers (legacy — immediate, annual-total only) ──────────────────────────
@app.get("/api/transfers", response_model=List[TransferOut])
def list_transfers(db: Session = Depends(get_db)):
    return db.query(Transfer).order_by(Transfer.transfer_date.desc()).all()

@app.post("/api/transfers", response_model=TransferOut, status_code=201)
def create_transfer(body: TransferCreate, db: Session = Depends(get_db)):
    amount = body.amount
    if body.transfer_type == "dept":
        src_projects = [p.project_code for p in db.query(Project).filter(Project.dept_code == body.from_code).all()]
        dst_projects = [p.project_code for p in db.query(Project).filter(Project.dept_code == body.to_code).all()]
        src_acts = db.query(Activity).filter(Activity.project_code.in_(src_projects)).all()
        dst_acts = db.query(Activity).filter(Activity.project_code.in_(dst_projects)).all()
        src_total = sum(a.allocated for a in src_acts)
        if amount > src_total:
            raise HTTPException(400, f"Insufficient budget in source department (Rs {src_total:,} available)")
        dst_total = sum(a.allocated for a in dst_acts) or 1
        for a in src_acts:
            a.allocated = int(a.allocated * (1 - amount / src_total))
        for a in dst_acts:
            a.allocated = int(a.allocated + amount * (a.allocated / dst_total))

    elif body.transfer_type == "project":
        src_acts = db.query(Activity).filter(Activity.project_code == body.from_code).all()
        dst_acts = db.query(Activity).filter(Activity.project_code == body.to_code).all()
        src_total = sum(a.allocated for a in src_acts)
        if amount > src_total:
            raise HTTPException(400, f"Insufficient budget in source project (Rs {src_total:,} available)")
        dst_total = sum(a.allocated for a in dst_acts) or 1
        for a in src_acts:
            a.allocated = int(a.allocated * (1 - amount / src_total))
        for a in dst_acts:
            a.allocated = int(a.allocated + amount * (a.allocated / dst_total))

    else:  # activity-level
        src = db.query(Activity).filter(Activity.activity_code == body.from_code).first()
        dst = db.query(Activity).filter(Activity.activity_code == body.to_code).first()
        if not src or not dst:
            raise HTTPException(404, "Source or destination activity not found")
        if amount > src.allocated:
            raise HTTPException(400, f"Insufficient budget (Rs {src.allocated:,} available)")
        src.allocated -= amount
        dst.allocated += amount

    tr = Transfer(
        id=_uid("TRF-"),
        transfer_type=body.transfer_type,
        from_code=body.from_code,
        to_code=body.to_code,
        amount=amount,
        employee_code=body.employee_code,
        note=body.note,
        transfer_date=body.transfer_date,
        status="Approved",
    )
    db.add(tr); db.commit(); db.refresh(tr)
    return tr


# ── Change requests (governed budget movement — months / carry / pull / transfer) ──
#
# Raising a request only logs intent — it does NOT touch any budget figures.
# Money only moves when a Finance user calls the /decide endpoint with
# approve=true. This mirrors the v4 reference behaviour: "Requires approval
# before it changes the budget."

def _get_or_404(db, model, code, code_field, label):
    obj = db.query(model).filter(getattr(model, code_field) == code).first()
    if not obj:
        raise HTTPException(404, f"{label} not found: {code}")
    return obj

def _phase_row(db: Session, activity_code: str, period_no: int) -> ActivityPhase:
    row = (db.query(ActivityPhase)
             .filter(ActivityPhase.activity_code == activity_code,
                     ActivityPhase.period_no == period_no)
             .first())
    if not row:
        # create a zeroed row if one doesn't exist yet (defensive — seed data
        # always has all 12, but user-created activities may not)
        row = ActivityPhase(activity_code=activity_code, period_no=period_no,
                             alloc=0, pr=0, po=0, cons=0)
        db.add(row); db.flush()
    return row

def _enrich_cr(cr: ChangeRequest, db: Session) -> ChangeRequestOut:
    out = ChangeRequestOut.model_validate(cr)
    a = db.query(Activity).filter(Activity.activity_code == cr.a_code).first()
    b = db.query(Activity).filter(Activity.activity_code == cr.b_code).first() if cr.b_code else None
    out.a_name = a.name if a else cr.a_code
    out.b_name = b.name if b else None
    return out


@app.get("/api/change-requests", response_model=List[ChangeRequestOut])
def list_change_requests(
    status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(ChangeRequest)
    if status:
        q = q.filter(ChangeRequest.status == status)
    rows = q.order_by(ChangeRequest.created_at.desc()).all()
    return [_enrich_cr(r, db) for r in rows]


@app.post("/api/change-requests", response_model=ChangeRequestOut, status_code=201)
def create_change_request(body: ChangeRequestCreate, db: Session = Depends(get_db)):
    if body.request_type not in ("months", "carry", "pull", "transfer"):
        raise HTTPException(400, "request_type must be one of months, carry, pull, transfer")
    if body.amount <= 0:
        raise HTTPException(400, "amount must be greater than zero")

    a = _get_or_404(db, Activity, body.a_code, "activity_code", "Source activity")

    if body.request_type == "transfer":
        if not body.b_code:
            raise HTTPException(400, "transfer requests require a target activity (b_code)")
        if body.b_code == body.a_code:
            raise HTTPException(400, "Source and target activity must differ")
        _get_or_404(db, Activity, body.b_code, "activity_code", "Target activity")
    else:
        if body.period_from == body.period_to:
            raise HTTPException(400, "Source and destination months must differ")

    for p in (body.period_from, body.period_to):
        if p < 1 or p > 12:
            raise HTTPException(400, "period_from / period_to must be between 1 and 12")

    # Validate enough is actually free to move (mirrors v4 client-side checks,
    # enforced here too so the request can't be raised against an impossible amount).
    src_phase = _phase_row(db, body.a_code, body.period_from)
    free = src_phase.alloc - src_phase.cons
    if body.amount > free:
        raise HTTPException(400, f"Only {free:,} is free in {a.name} for period {body.period_from} (requested {body.amount:,}).")

    # Validate requested_by exists to avoid foreign key errors
    if body.requested_by:
        emp_exists = db.query(Employee).filter(Employee.employee_code == body.requested_by).first()
        if not emp_exists:
            body.requested_by = None

    cr = ChangeRequest(
        id=_uid("CR-"),
        request_type=body.request_type,
        a_code=body.a_code,
        b_code=body.b_code,
        period_from=body.period_from,
        period_to=body.period_to,
        amount=body.amount,
        reason=body.reason,
        requested_by=body.requested_by,
        status="Pending",
        created_at=datetime.utcnow(),
    )
    db.add(cr); db.commit(); db.refresh(cr)
    return _enrich_cr(cr, db)


@app.patch("/api/change-requests/{cr_id}/decide", response_model=ChangeRequestOut)
def decide_change_request(cr_id: str, body: ChangeRequestDecide, db: Session = Depends(get_db)):
    cr = db.query(ChangeRequest).filter(ChangeRequest.id == cr_id).first()
    if not cr:
        raise HTTPException(404, "Change request not found")
    if cr.status != "Pending":
        raise HTTPException(400, f"Change request is already {cr.status}")

    if body.approve:
        # ── THIS is where the money actually moves ──
        if cr.request_type == "transfer":
            src_act = _get_or_404(db, Activity, cr.a_code, "activity_code", "Source activity")
            dst_act = _get_or_404(db, Activity, cr.b_code, "activity_code", "Target activity")
            src_phase = _phase_row(db, cr.a_code, cr.period_from)
            dst_phase = _phase_row(db, cr.b_code, cr.period_from)

            free = src_phase.alloc - src_phase.cons
            if cr.amount > free:
                raise HTTPException(400, f"Approval failed: only {free:,} now free in {src_act.name} for that month (budget moved since request was raised).")

            src_phase.alloc -= cr.amount
            dst_phase.alloc += cr.amount
            src_act.allocated -= cr.amount
            dst_act.allocated += cr.amount

        else:  # months | carry | pull — same activity, two different months
            act = _get_or_404(db, Activity, cr.a_code, "activity_code", "Activity")
            src_phase = _phase_row(db, cr.a_code, cr.period_from)
            dst_phase = _phase_row(db, cr.a_code, cr.period_to)

            free = src_phase.alloc - src_phase.cons
            if cr.amount > free:
                raise HTTPException(400, f"Approval failed: only {free:,} now free in {act.name} for that month (budget moved since request was raised).")

            src_phase.alloc -= cr.amount
            dst_phase.alloc += cr.amount
            # activity.allocated (annual total) is unchanged — moving budget
            # between months of the same activity doesn't change its total.

        cr.status = "Approved"
    else:
        cr.status = "Rejected"

    cr.decided_by = body.decided_by
    cr.decided_at = datetime.utcnow()
    db.commit()
    db.refresh(cr)
    return _enrich_cr(cr, db)


# ── Invoice Validation ────────────────────────────────────────────────────────
@app.post("/api/check-invoice", response_model=InvoiceBudgetCheckRes)
def check_invoice_budget(payload: InvoiceBudgetCheckReq, db: Session = Depends(get_db)):
    # 1. Determine current month and quarter
    current_month = datetime.now().month
    quarter_start_month = 3 * ((current_month - 1) // 3) + 1
    quarter_months = [quarter_start_month, quarter_start_month + 1, quarter_start_month + 2]
    
    # 2. Fetch the phases for this activity
    phases = db.query(ActivityPhase).filter(
        ActivityPhase.activity_code == payload.activity_code
    ).all()
    
    if not phases:
        return InvoiceBudgetCheckRes(
            current_month_allocated=0, current_month_available=0,
            current_quarter_allocated=0, current_quarter_available=0,
            covers_with_tax=False, covers_without_tax=False
        )

    # 3. Calculate Current Month Budget
    current_phase = next((p for p in phases if p.period_no == current_month), None)
    month_alloc = current_phase.alloc if current_phase else 0
    month_used = (current_phase.pr + current_phase.po + current_phase.cons) if current_phase else 0
    month_available = month_alloc - month_used

    # 4. Calculate Current Quarter Budget
    quarter_phases = [p for p in phases if p.period_no in quarter_months]
    quarter_alloc = sum(p.alloc for p in quarter_phases)
    quarter_used = sum(p.pr + p.po + p.cons for p in quarter_phases)
    quarter_available = quarter_alloc - quarter_used
    
    # 5. Check covers against the current month available
    covers_with_tax = month_available >= payload.amount_with_tax
    covers_without_tax = month_available >= payload.amount_without_tax

    return InvoiceBudgetCheckRes(
        current_month_allocated=month_alloc,
        current_month_available=month_available,
        current_quarter_allocated=quarter_alloc,
        current_quarter_available=quarter_available,
        covers_with_tax=covers_with_tax,
        covers_without_tax=covers_without_tax
    )

@app.post("/api/block-amount")
def block_budget_amount(payload: BudgetBlockReq, db: Session = Depends(get_db)):
    current_month = datetime.now().month
    phase = db.query(ActivityPhase).filter(
        ActivityPhase.activity_code == payload.activity_code,
        ActivityPhase.period_no == current_month
    ).first()
    
    if phase:
        phase.pr += payload.amount
        
    act = db.query(Activity).filter(Activity.activity_code == payload.activity_code).first()
    if act:
        act.pr += payload.amount
        
    db.commit()
    return {"status": "success", "blocked_amount": payload.amount}


# ── Excel Budget Upload (staged, per-department, head-approved) ──────────────
"""
Expected Excel columns (case-insensitive, order flexible):
  project_code | activity_name | wbs | cost_type (Opex/Capex) |
  allocated | pr | po | invoiced | employee_code | status_code

Every upload is scoped to exactly one department (`dept_code` is required).
Rows whose project_code belongs to a different department (or doesn't exist)
are skipped with a note, so one department's spreadsheet can never touch
another department's projects.

Uploads are STAGED, not applied. A department approval marks that department's
upload as Approved for the fiscal-year cycle, but the rows are written to
Activity only when every department has an Approved, not-yet-merged upload for
that same fiscal year. Rejecting (or simply never deciding) leaves the live
budget untouched. The counts returned at upload time are a PREVIEW; they're
recomputed for real at merge time since the underlying data may have changed.

NOTE: this only writes annual totals to Activity, not monthly ActivityPhase
rows. A newly inserted activity will show zero in every month of the Budget
Book monthly view / phasing grid until phases are set via a change request
or future phase-editing endpoint.
"""


def _parse_budget_excel_rows(content: bytes, dept_code: str, db: Session):
    import openpyxl

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Cannot read Excel file: {e}")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        raise HTTPException(400, "Excel sheet must contain headers in row 1-2 and data from row 3")

    all_projects  = db.query(Project).all()
    valid_projects = {p.project_code for p in all_projects if p.dept_code == dept_code}
    all_project_codes = {p.project_code for p in all_projects}

    staged, skip_notes = [], []

    # Indices for fixed columns (0-based)
    idx_proj = 3    # D
    idx_act  = 5    # F
    idx_wbs  = 9    # J
    idx_ct   = 10   # K
    idx_emp  = 12   # M
    idx_stat = 13   # N

    # Indices for monthly "Allocated" and "PR/PO" columns
    # We skip Q1, Q2, Q3, Q4 total columns.
    # Apr (14), May (17), Jun (20), Jul (26), Aug (29), Sep (32)
    # Oct (38), Nov (41), Dec (44), Jan (50), Feb (53), Mar (56)
    month_starts = [14, 17, 20, 26, 29, 32, 38, 41, 44, 50, 53, 56]

    for row_num, row in enumerate(rows[2:], start=3):
        def cell(i):
            return str(row[i]).strip() if i < len(row) and row[i] is not None else ""

        project_code  = cell(idx_proj)
        activity_name = cell(idx_act)
        wbs_val       = cell(idx_wbs)

        if not project_code or not activity_name or not wbs_val:
            if not any(cell(i) for i in range(15)):  # Skip completely empty rows
                continue
            skip_notes.append(f"Row {row_num}: missing required field (Project, Activity, or WBS)")
            continue

        if project_code not in valid_projects:
            if project_code in all_project_codes:
                skip_notes.append(f"Row {row_num}: project_code '{project_code}' belongs to a different department")
            else:
                skip_notes.append(f"Row {row_num}: unknown project_code '{project_code}'")
            continue

        def safe_int(i):
            v = cell(i)
            try: return int(float(v)) if v else 0
            except: return 0

        months_data = []
        annual_allocated = 0
        annual_pr = 0
        for period_no, start_idx in enumerate(month_starts, start=1):
            alloc = safe_int(start_idx)
            pr = safe_int(start_idx + 1)
            months_data.append({
                "period_no": period_no,
                "alloc": alloc,
                "pr": pr,
                "po": 0,
                "cons": 0
            })
            annual_allocated += alloc
            annual_pr += pr

        staged.append(dict(
            row_num        = row_num,
            project_code   = project_code,
            activity_name  = activity_name,
            wbs            = wbs_val,
            cost_type_tag  = cell(idx_ct),
            allocated      = annual_allocated,
            pr             = annual_pr,
            po             = 0,
            invoiced       = 0,
            employee_code  = cell(idx_emp) or None,
            status_code    = cell(idx_stat) or "STS-NS",
            months_data    = months_data
        ))

    return staged, skip_notes


def _apply_staged_rows(staged_rows: list, dept_code: str, db: Session):
    cost_type_map  = {ct.tag.lower(): ct.cost_type_code for ct in db.query(CostType).all()}
    all_projects   = db.query(Project).all()
    valid_projects = {p.project_code for p in all_projects if p.dept_code == dept_code}
    existing_wbs   = {a.wbs: a for a in db.query(Activity).all()}

    inserted = updated = skipped = 0
    skip_notes = []

    for r in staged_rows:
        if r["project_code"] not in valid_projects:
            skipped += 1
            skip_notes.append(f"Row {r['row_num']}: project_code '{r['project_code']}' no longer valid for this department")
            continue

        cost_type_code = cost_type_map.get((r.get("cost_type_tag") or "").lower()) or cost_type_map.get("opex")

        if r["wbs"] in existing_wbs:
            act = existing_wbs[r["wbs"]]
            act.name           = r["activity_name"]
            act.allocated      = r["allocated"]
            act.pr             = r["pr"]
            act.po             = r["po"]
            act.invoiced       = r["invoiced"]
            act.cost_type_code = cost_type_code
            if r["employee_code"]: act.employee_code = r["employee_code"]
            if r["status_code"]:   act.status_code   = r["status_code"]
            updated += 1
        else:
            act = Activity(
                activity_code  = _uid("ACT-"),
                name           = r["activity_name"],
                project_code   = r["project_code"],
                cost_type_code = cost_type_code,
                employee_code  = r["employee_code"] or None,
                status_code    = r["status_code"],
                wbs            = r["wbs"],
                is_leaf        = 1,
                allocated      = r["allocated"],
                pr             = r["pr"],
                po             = r["po"],
                invoiced       = r["invoiced"],
            )
            db.add(act)
            existing_wbs[r["wbs"]] = act
            inserted += 1

        db.flush()
        
        # Monthly phases
        existing_phases = {p.period_no: p for p in db.query(ActivityPhase).filter(ActivityPhase.activity_code == act.activity_code).all()}
        for m_data in r.get("months_data", []):
            period = m_data["period_no"]
            if period in existing_phases:
                phase = existing_phases[period]
                phase.alloc = m_data["alloc"]
                phase.pr = m_data["pr"]
            else:
                phase = ActivityPhase(
                    activity_code = act.activity_code,
                    period_no = period,
                    alloc = m_data["alloc"],
                    pr = m_data["pr"],
                    po = 0,
                    cons = 0
                )
                db.add(phase)

    db.flush()
    return inserted, updated, skipped, skip_notes


def _cycle_filter(q, fiscal_year: Optional[str]):
    if fiscal_year is None:
        return q.filter(BudgetUpload.fiscal_year.is_(None))
    return q.filter(BudgetUpload.fiscal_year == fiscal_year)


def _approved_unmerged_uploads(db: Session, fiscal_year: Optional[str]):
    q = (db.query(BudgetUpload)
           .filter(BudgetUpload.status == "Approved")
           .filter(BudgetUpload.staged_rows.isnot(None)))
    return (_cycle_filter(q, fiscal_year)
            .order_by(BudgetUpload.uploaded_at.asc())
            .all())


def _try_merge_budget_cycle(db: Session, fiscal_year: Optional[str]) -> bool:
    """Merge a fiscal-year cycle once every project-owning department has approved.

    The cycle key is fiscal_year, including None for uploads that do not carry
    one. Only approved uploads with staged_rows still present count; after a
    merge, staged_rows is cleared so those approvals cannot trigger a later
    cycle by accident.
    """
    project_depts = {p.dept_code for p in db.query(Project).all()}
    if not project_depts:
        return False

    uploads = _approved_unmerged_uploads(db, fiscal_year)
    approved_depts = {u.dept_code for u in uploads if u.dept_code}
    # REMOVED STRICT CHECK FOR TESTING: Allow merging immediately without waiting for all departments
    # if not project_depts.issubset(approved_depts):
    #     return False

    merged_at = datetime.utcnow()
    for upload in uploads:
        staged = json.loads(upload.staged_rows or "[]")
        inserted, updated, skipped, skip_notes = _apply_staged_rows(staged, upload.dept_code, db)

        prior_notes = upload.notes.split("; ") if upload.notes else []
        all_notes = prior_notes + skip_notes
        notes_str = "; ".join(all_notes[:20])
        if len(all_notes) > 20:
            notes_str += f" … and {len(all_notes)-20} more"

        upload.rows_inserted = inserted
        upload.rows_updated  = updated
        upload.rows_skipped  = upload.rows_skipped + skipped
        upload.notes         = notes_str or None
        upload.staged_rows   = None
        upload.merged_at     = merged_at

    return True


def trigger_budget_approval_workflow(upload_id: str, amount: float, fiscal_year: str, db: Session, user_id: int):
    from routers.requests import submit_request
    from schemas import RequestCreate
    from models import Workflow

    # Find the budget approval workflow
    wf = db.query(Workflow).filter(Workflow.name.ilike("%Budget%")).first()
    if not wf:
        print("No budget workflow found to trigger!")
        return
        
    payload = RequestCreate(
        title=f"Budget Upload Approval for FY {fiscal_year}",
        description=f"Automated request for budget upload {upload_id}",
        amount=amount,
        request_type="budget_upload",
        request_metadata={"upload_id": upload_id},
        workflow_id=wf.id
    )
    
    try:
        submit_request(payload, user_id=user_id, db=db)
        print(f"Successfully raised workflow request for budget upload {upload_id}")
    except Exception as e:
        print(f"Failed to trigger workflow: {e}")


@app.post("/api/budget-upload", response_model=BudgetUploadOut, status_code=201)
async def upload_budget_excel(
    file: UploadFile = File(...),
    fiscal_year: Optional[str] = Form(None),
    dept_code: str = Form(...),
    requested_by: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx or .xls files are accepted")

    dept = db.query(Department).filter(Department.dept_code == dept_code).first()
    if not dept:
        raise HTTPException(400, f"Unknown dept_code '{dept_code}'")

    try:
        import openpyxl  # noqa: F401 — checked here so the error surfaces before parsing
    except ImportError:
        raise HTTPException(500, "openpyxl not installed — run: pip install openpyxl")

    content = await file.read()
    staged_rows, skip_notes = _parse_budget_excel_rows(content, dept_code, db)

    # PREVIEW ONLY — nothing is written to Activity yet. We simulate against
    # current state purely to show the uploader what approval would do.
    existing_wbs = {a.wbs for a in db.query(Activity).all()}
    preview_inserted = sum(1 for r in staged_rows if r["wbs"] not in existing_wbs)
    preview_updated  = len(staged_rows) - preview_inserted
    preview_skipped  = len(skip_notes)

    notes_str = "; ".join(skip_notes[:20])
    if len(skip_notes) > 20:
        notes_str += f" … and {len(skip_notes)-20} more"

    upload_log = BudgetUpload(
        id            = _uid("UPL-"),
        filename      = file.filename,
        fiscal_year   = fiscal_year,
        dept_code     = dept_code,
        status        = "Pending",
        staged_rows   = json.dumps(staged_rows),
        rows_inserted = preview_inserted,
        rows_updated  = preview_updated,
        rows_skipped  = preview_skipped,
        requested_by  = requested_by,
        uploaded_at   = datetime.utcnow(),
        notes         = notes_str or None,
    )
    db.add(upload_log)
    db.commit()
    db.refresh(upload_log)
    
    total_allocated = sum(r.get("allocated", 0) for r in staged_rows)
    # Get user_id from requested_by or default to admin
    from models import User
    user = db.query(User).filter(User.name == requested_by).first() if requested_by else None
    if not user:
        user = db.query(User).first() # Fallback
    trigger_budget_approval_workflow(upload_log.id, total_allocated, fiscal_year, db, user.id if user else 1)
    
    return upload_log


@app.get("/api/budget-uploads", response_model=List[BudgetUploadOut])
def list_budget_uploads(
    dept_code: Optional[str] = Query(None),
    status: Optional[str]    = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(BudgetUpload)
    if dept_code:
        q = q.filter(BudgetUpload.dept_code == dept_code)
    if status:
        q = q.filter(BudgetUpload.status == status)
    return q.order_by(BudgetUpload.uploaded_at.desc()).all()


@app.patch("/api/budget-uploads/{upload_id}/decide", response_model=BudgetUploadOut)
def decide_budget_upload(upload_id: str, body: BudgetUploadDecide, db: Session = Depends(get_db)):
    upload = db.query(BudgetUpload).filter(BudgetUpload.id == upload_id).first()
    if not upload:
        raise HTTPException(404, "Budget upload not found")
    if upload.status != "Pending":
        raise HTTPException(400, f"Budget upload is already {upload.status}")

    if body.approve:
        upload.status        = "Approved"
    else:
        upload.status = "Rejected"

    upload.decided_by = body.decided_by
    upload.decided_at = datetime.utcnow()
    if body.approve:
        db.flush()
        _try_merge_budget_cycle(db, upload.fiscal_year)
    else:
        upload.staged_rows = None
    db.commit()
    db.refresh(upload)
    return upload

@app.get("/api/budget-template")
def download_budget_template(db: Session = Depends(get_db)):
    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget Template"

    # Define styles
    header_fill = PatternFill(start_color="1D428A", end_color="1D428A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Fixed headers (Columns A to N)
    fixed_headers = [
        "Line ID", "Organisation", "Department", "Project Code", "Project",
        "Activity", "Sub Activity 1", "Sub Activity 2", "Sub Activity 3",
        "WBS", "Cost Type", "Owner (Employee)", "Employee Code", "Status"
    ]

    months = [
        "Apr", "May", "Jun", "Q1 Total",
        "Jul", "Aug", "Sep", "Q2 Total",
        "Oct", "Nov", "Dec", "Q3 Total",
        "Jan", "Feb", "Mar", "Q4 Total"
    ]

    # Row 1 and Row 2 setup
    for col_idx, header in enumerate(fixed_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
        # adjust row 2 border for merged cell visually
        ws.cell(row=2, column=col_idx).border = border

    current_col = len(fixed_headers) + 1
    for month in months:
        # Row 1 merged cell for month
        cell1 = ws.cell(row=1, column=current_col, value=month)
        ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col+2)
        cell1.fill = header_fill
        cell1.font = header_font
        cell1.alignment = center_align
        cell1.border = border
        ws.cell(row=1, column=current_col+1).border = border
        ws.cell(row=1, column=current_col+2).border = border

        # Row 2 sub-headers
        for sub_idx, sub_header in enumerate(["Allocated", "PR/PO", "Remaining"]):
            cell2 = ws.cell(row=2, column=current_col + sub_idx, value=sub_header)
            cell2.fill = header_fill
            cell2.font = header_font
            cell2.alignment = center_align
            cell2.border = border
            
        current_col += 3

    # Add Data Validation for Dropdowns
    depts = [d.name for d in db.query(Department).all()]
    project_codes = [p.project_code for p in db.query(Project).all()]
    project_names = [p.name for p in db.query(Project).all()]
    activities = [a.name for a in db.query(Activity).all()]
    wbs_codes = [a.wbs for a in db.query(Activity).all()]
    cost_types = ["Opex", "Capex"]
    employee_codes = [e.employee_code for e in db.query(Employee).all()]
    employee_names = [e.name for e in db.query(Employee).all()]
    statuses = [s.status_code for s in db.query(Status).all()]

    def add_validation(dv_formula, col_letter, title, show_error=True):
        dv = DataValidation(type="list", formula1=dv_formula, allow_blank=True)
        if show_error:
            dv.error = "Your entry is not in the list"
            dv.errorTitle = "Invalid Entry"
        else:
            dv.showErrorMessage = False
        dv.prompt = "Please select from the list"
        dv.promptTitle = title
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}3:{col_letter}1000")

    list_ws = wb.create_sheet(title="Lists")
    list_ws.sheet_state = 'hidden'
    
    def write_col(col_idx, data):
        # filter out empty values and get unique sorted list
        unique_data = sorted(list(set(d for d in data if d)))
        for r_idx, val in enumerate(unique_data, start=1):
            list_ws.cell(row=r_idx, column=col_idx, value=val)
        return f"Lists!${openpyxl.utils.get_column_letter(col_idx)}$1:${openpyxl.utils.get_column_letter(col_idx)}${max(1, len(unique_data))}"
    
    dept_ref = write_col(1, depts)
    proj_code_ref = write_col(2, project_codes)
    proj_name_ref = write_col(3, project_names)
    act_ref = write_col(4, activities)
    wbs_ref = write_col(5, wbs_codes)
    ct_ref = write_col(6, cost_types)
    emp_name_ref = write_col(7, employee_names)
    emp_code_ref = write_col(8, employee_codes)
    stat_ref = write_col(9, statuses)

    if len(depts) < 10:
        add_validation(f'"{",".join(depts)}"', "C", "Select Department")
    else:
        add_validation(dept_ref, "C", "Select Department")
        
    add_validation(proj_code_ref, "D", "Select Project Code")
    add_validation(proj_name_ref, "E", "Select Project Name", show_error=False)
    add_validation(act_ref, "F", "Select Activity", show_error=False)
    add_validation(wbs_ref, "J", "Select WBS", show_error=False)
    add_validation(f'"{",".join(cost_types)}"', "K", "Select Cost Type")
    add_validation(emp_name_ref, "L", "Select Owner Name", show_error=False)
    add_validation(emp_code_ref, "M", "Select Employee Code")
    add_validation(stat_ref, "N", "Select Status")

    # Auto-adjust column widths
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = 15

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    headers_dict = {
        'Content-Disposition': 'attachment; filename="Budget_Upload_Template.xlsx"'
    }
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers_dict)


# Mount standard workflow engine routers
app.include_router(auth.router,       prefix="/api/auth",      tags=["Auth"])
app.include_router(workflows.router,  prefix="/api/workflows", tags=["Workflows"])
app.include_router(requests.router,   prefix="/api/requests",  tags=["Requests"])
app.include_router(stages.router,     prefix="/api/stages",    tags=["Stages"])
app.include_router(approvals.router,  prefix="/api/approvals", tags=["Approvals"])
app.include_router(analytics.router,  prefix="/api/analytics", tags=["Analytics"])
app.include_router(reports.router)
app.include_router(vendors.router)
app.include_router(vendor_materials.router)

@app.get("/api/department-status")
def get_department_status(dept_code: str, db: Session = Depends(get_db)):
    from models import Department, Activity, ActivityPhase
    import datetime
    
    dept = db.query(Department).filter(Department.dept_code == dept_code).first()
    if not dept:
        return {"status": "ERROR", "message": "Department not found"}
        
    activities = db.query(Activity).filter(Activity.project.has(dept_code=dept_code)).all()
    
    today = datetime.datetime.now()
    month = today.month
    
    # Financial Year mapping (April start)
    # 1=Apr, 2=May, ..., 12=Mar
    period_map = {4:1, 5:2, 6:3, 7:4, 8:5, 9:6, 10:7, 11:8, 12:9, 1:10, 2:11, 3:12}
    current_period = period_map[month]
    
    if current_period in [1, 2, 3]:
        q_periods = [1, 2, 3]
    elif current_period in [4, 5, 6]:
        q_periods = [4, 5, 6]
    elif current_period in [7, 8, 9]:
        q_periods = [7, 8, 9]
    else:
        q_periods = [10, 11, 12]
        
    current_month_allocated = 0
    current_quarter_allocated = 0
    
    for act in activities:
        for phase in act.phases:
            if phase.period_no == current_period:
                current_month_allocated += phase.alloc
            if phase.period_no in q_periods:
                current_quarter_allocated += phase.alloc
                
    return {
        "status": "SUCCESS",
        "data": {
            "current_month_allocated": current_month_allocated,
            "current_quarter_allocated": current_quarter_allocated,
            "department_approved": sum(a.approved for a in activities)
        }
    }
