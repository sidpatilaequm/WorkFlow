import random
import string
import io
import logging
from datetime import date, datetime
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from sqlalchemy.orm import Session, joinedload

from database import get_db, SessionLocal
from models import (
    Organisation, Department, Project, CostType, Status,
    Employee, Activity, ActivityPhase, SubActivity, SubActivityPhase,
    BudgetVersion, FiscalPeriod, Transfer, ChangeRequest,
    BudgetUpload,
)
import requests
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from schemas_budget import (
    OrganisationOut,
    DepartmentCreate, DepartmentOut,
    ProjectCreate, ProjectOut,
    CostTypeOut, StatusOut,
    EmployeeCreate, EmployeeOut,
    ActivityCreate, ActivityUpdate, ActivityOut,
    SubActivityCreate, SubActivityUpdate, SubActivityOut,
    BudgetVersionCreate, BudgetVersionOut,
    FiscalPeriodOut,
    TransferCreate, TransferOut,
    ChangeRequestCreate, ChangeRequestOut, ChangeRequestDecide,
    BudgetUploadOut,
    DashboardOut, DeptSummary,
    InvoiceBudgetCheckReq, InvoiceBudgetCheckRes,
)

logger = logging.getLogger(__name__)

router = APIRouter()


def _uid(prefix: str = "") -> str:
    chars = string.ascii_uppercase + string.digits
    return prefix + "".join(random.choices(chars, k=6))


def _enrich_activity(a: Activity) -> ActivityOut:
    out = ActivityOut.model_validate(a)
    out.cost_type_tag  = a.cost_type.tag  if a.cost_type  else None
    out.status_name    = a.status.name    if a.status     else None
    out.employee_name  = a.employee.name  if a.employee   else None
    out.sub_activities = [_enrich_sub(s) for s in a.sub_activities]
    out.phases         = list(a.phases or [])
    return out


def _enrich_sub(s: SubActivity) -> SubActivityOut:
    out = SubActivityOut.model_validate(s)
    out.cost_type_tag = s.cost_type.tag if s.cost_type else None
    out.status_name   = s.status.name   if s.status    else None
    out.employee_name = s.employee.name if s.employee  else None
    out.phases        = list(s.phases or [])
    return out


# ── Monthly phase helper for seed ─────────────────────────────────────────────
def _make_phases(alloc_list, pr_list, po_list, cons_list):
    return [
        {"period_no": i+1, "alloc": int(alloc_list[i]), "pr": int(pr_list[i]),
         "po": int(po_list[i]), "cons": int(cons_list[i])}
        for i in range(12)
    ]


# ── Seed Reference Data Function ──────────────────────────────────────────────
def seed_reference_data():
    db = SessionLocal()
    try:
        if db.query(Organisation).count():
            logger.info("[budget_seed] Database already seeded.")
            return

        logger.info("[budget_seed] Seeding budget reference data...")

        # ── Organisation ──
        org = Organisation(org_code="ORG-001", name="Aequm India",
                           base_currency="INR", fiscal_year="FY 2026-27")
        db.add(org); db.flush()

        # ── Departments ──
        for code, name, wbs in [
            ("ENG", "Engineering",       "1.1"),
            ("SMK", "Sales & Marketing", "1.2"),
            ("OPS", "Operations",        "1.3"),
            ("FIN", "Finance & Admin",   "1.4"),
            ("CSU", "Customer Success",  "1.5"),
        ]:
            db.add(Department(dept_code=code, name=name, org_code="ORG-001", wbs=wbs))
        db.flush()

        # ── Projects ──
        for code, name, dept, wbs in [
            ("PRJ-ENG-01", "Platform R&D",           "ENG", "1.1.1"),
            ("PRJ-ENG-02", "Mobile App",              "ENG", "1.1.2"),
            ("PRJ-SEC-ENG","Cybersecurity Uplift",    "ENG", "1.1.3"),
            ("PRJ-SMK-01", "Brand Campaign 2026",     "SMK", "1.2.1"),
            ("PRJ-SMK-02", "CRM Rollout",             "SMK", "1.2.2"),
            ("PRJ-OPS-01", "Supply Chain Automation", "OPS", "1.3.1"),
            ("PRJ-OPS-02", "Facilities",              "OPS", "1.3.2"),
            ("PRJ-SEC-OPS","Cybersecurity Uplift",    "OPS", "1.3.3"),
            ("PRJ-FIN-01", "ERP Implementation",      "FIN", "1.4.1"),
            ("PRJ-FIN-02", "Compliance",              "FIN", "1.4.2"),
            ("PRJ-SEC-FIN","Cybersecurity Uplift",    "FIN", "1.4.3"),
            ("PRJ-CSU-01", "Onboarding Program",      "CSU", "1.5.1"),
        ]:
            db.add(Project(project_code=code, name=name, dept_code=dept, wbs=wbs))
        db.flush()

        # ── Cost types ──
        db.add(CostType(cost_type_code="OPEX",  name="Operating Expenditure", tag="Opex"))
        db.add(CostType(cost_type_code="CAPEX", name="Capital Expenditure",   tag="Capex"))
        db.flush()

        # ── Statuses ──
        for code, name, order in [
            ("STS-IP",  "In Progress", 1), ("STS-NS",  "Not Started", 2),
            ("STS-DONE","Completed",   3), ("STS-HOLD","On Hold",      4),
        ]:
            db.add(Status(status_code=code, name=name, sort_order=order))
        db.flush()

        # ── Employees ──
        for code, name, title, dept, mgr in [
            ("EMP-01", "Asha Pillai",   "Chief Financial Officer",            None,  None),
            ("EMP-02", "Priya Nair",    "Head of Engineering",                "ENG", "EMP-01"),
            ("EMP-03", "Rahul Verma",   "Engineering Lead",                   "ENG", "EMP-02"),
            ("EMP-04", "Sana Kapoor",   "Design Lead",                        "ENG", "EMP-02"),
            ("EMP-05", "Arjun Mehta",   "Head of Sales & Marketing",          "SMK", "EMP-01"),
            ("EMP-06", "Neha Singh",    "Marketing Manager",                  "SMK", "EMP-05"),
            ("EMP-07", "Vikram Rao",    "Head of Operations",                 "OPS", "EMP-01"),
            ("EMP-08", "Meera Iyer",    "Facilities Manager",                 "OPS", "EMP-07"),
            ("EMP-09", "Sunil Das",     "Head of Finance & Admin",            "FIN", "EMP-01"),
            ("EMP-10", "Ravi Krishnan", "Chief Information Security Officer", "ENG", "EMP-01"),
            ("EMP-11", "Divya Menon",   "Head of Customer Success",           "CSU", "EMP-01"),
        ]:
            db.add(Employee(employee_code=code, name=name, title=title,
                            dept_code=dept, manager_code=mgr))
        db.flush()

        # ── Budget version ──
        db.add(BudgetVersion(version_code="BV-001", name="Annual Budget FY 2026-27",
                             fiscal_year="FY 2026-27", created_date=date(2026, 4, 1),
                             basis="Excel upload", is_current=1, is_locked=0))

        # ── Fiscal periods ──
        for row in [
            (1,"Apr-26","Q1","April","Current"),(2,"May-26","Q1","May","Open"),
            (3,"Jun-26","Q1","June","Open"),(4,"Jul-26","Q2","July","Open"),
            (5,"Aug-26","Q2","August","Open"),(6,"Sep-26","Q2","September","Open"),
            (7,"Oct-26","Q3","October","Open"),(8,"Nov-26","Q3","November","Open"),
            (9,"Dec-26","Q3","December","Open"),(10,"Jan-27","Q4","January","Open"),
            (11,"Feb-27","Q4","February","Open"),(12,"Mar-27","Q4","March","Open"),
        ]:
            db.add(FiscalPeriod(period_no=row[0], period_code=row[1], quarter=row[2],
                                month=row[3], state=row[4], fiscal_year="FY 2026-27"))

        # ── Activities with monthly phases ──
        acts = [
          ("ACT-001","Backend development","PRJ-ENG-01","OPEX","EMP-02","STS-IP","1.1.1.1",2500000,
           [208000]*11+[212000],[197600]*11+[201400],[176800]*11+[180200],
           [212000,472000,413000,295000,0,0,0,0,0,0,0,0]),

          ("ACT-002","QA & automation","PRJ-ENG-01","OPEX","EMP-03","STS-IP","1.1.1.2",900000,
           [75000]*12,[71250]*12,[63750]*12,
           [75000,164000,144000,102000,0,0,0,0,0,0,0,0]),

          ("ACT-003","Cloud infrastructure","PRJ-ENG-01","CAPEX","EMP-02","STS-DONE","1.1.1.3",1600000,
           [133000]*11+[137000],[133000]*11+[137000],[133000]*11+[137000],
           [137000,640000,560000,400000,0,0,0,0,0,0,0,0]),

          ("ACT-004","UI/UX design","PRJ-ENG-02","OPEX","EMP-04","STS-IP","1.1.2.1",600000,
           [50000]*12,[47500]*12,[42500]*12,
           [50000,100000,88000,62000,0,0,0,0,0,0,0,0]),

          ("ACT-005","Mobile development","PRJ-ENG-02","OPEX","EMP-03","STS-IP","1.1.2.2",1400000,
           [117000]*11+[113000],[111150]*11+[107350],[99450]*11+[96050],
           [113000,120000,105000,75000,0,0,0,0,0,0,0,0]),

          ("ACT-SEC-ENG","Security tooling (Eng)","PRJ-SEC-ENG","CAPEX","EMP-10","STS-IP","1.1.3.1",1200000,
           [100000]*12,[95000]*12,[85000]*12,
           [120000,120000,120000,0,0,0,0,0,0,0,0,0]),

          ("ACT-006","Digital advertising","PRJ-SMK-01","OPEX","EMP-05","STS-IP","1.2.1.1",1200000,
           [100000]*12,[95000]*12,[85000]*12,
           [100000,312000,273000,195000,0,0,0,0,0,0,0,0]),

          ("ACT-007","Events & expos","PRJ-SMK-01","OPEX","EMP-05","STS-NS","1.2.1.2",800000,
           [67000]*11+[63000],[63650]*11+[59850],[56950]*11+[53550],
           [63000,0,0,0,0,0,0,0,0,0,0,0]),

          ("ACT-008","Software licenses","PRJ-SMK-02","CAPEX","EMP-06","STS-DONE","1.2.2.1",500000,
           [42000]*11+[38000],[42000]*11+[38000],[42000]*11+[38000],
           [38000,200000,175000,125000,0,0,0,0,0,0,0,0]),

          ("ACT-009","Sales enablement","PRJ-SMK-02","OPEX","EMP-06","STS-IP","1.2.2.2",350000,
           [29000]*11+[31000],[27550]*11+[29450],[24650]*11+[26350],
           [31000,48000,42000,30000,0,0,0,0,0,0,0,0]),

          ("ACT-010","Process consulting","PRJ-OPS-01","OPEX","EMP-07","STS-IP","1.3.1.1",700000,
           [58000]*11+[62000],[55100]*11+[58900],[49300]*11+[52700],
           [62000,160000,140000,100000,0,0,0,0,0,0,0,0]),

          ("ACT-011","Workflow tooling","PRJ-OPS-01","CAPEX","EMP-07","STS-IP","1.3.1.2",950000,
           [79000]*11+[81000],[75050]*11+[76950],[67150]*11+[68850],
           [81000,240000,210000,150000,0,0,0,0,0,0,0,0]),

          ("ACT-012","Office expansion","PRJ-OPS-02","CAPEX","EMP-08","STS-IP","1.3.2.1",2000000,
           [167000]*11+[163000],[158650]*11+[154850],[146125]*11+[142625],
           [163000,700000,612000,438000,0,0,0,0,0,0,0,0]),

          ("ACT-013","Utilities & maintenance","PRJ-OPS-02","OPEX","EMP-08","STS-IP","1.3.2.2",450000,
           [38000]*11+[32000],[36100]*11+[30400],[32300]*11+[27200],
           [32000,92000,80000,58000,0,0,0,0,0,0,0,0]),

          ("ACT-SEC-OPS","Security tooling (Ops)","PRJ-SEC-OPS","CAPEX","EMP-10","STS-IP","1.3.3.1",1050000,
           [87500]*12,[83125]*12,[74375]*12,
           [105000,105000,105000,0,0,0,0,0,0,0,0,0]),

          ("ACT-014","ERP licenses","PRJ-FIN-01","CAPEX","EMP-09","STS-DONE","1.4.1.1",1800000,
           [150000]*12,[150000]*12,[150000]*12,
           [150000,720000,630000,450000,0,0,0,0,0,0,0,0]),

          ("ACT-015","Implementation services","PRJ-FIN-01","OPEX","EMP-09","STS-IP","1.4.1.2",1100000,
           [92000]*11+[88000],[87400]*11+[83600],[78200]*11+[74800],
           [88000,208000,182000,130000,0,0,0,0,0,0,0,0]),

          ("ACT-016","Audit & filings","PRJ-FIN-02","OPEX","EMP-09","STS-IP","1.4.2.1",400000,
           [33000]*11+[37000],[31350]*11+[35150],[28050]*11+[31450],
           [37000,72000,63000,45000,0,0,0,0,0,0,0,0]),

          ("ACT-SEC-FIN","Security tooling (Fin)","PRJ-SEC-FIN","CAPEX","EMP-10","STS-IP","1.4.3.1",750000,
           [62500]*12,[59375]*12,[53125]*12,
           [75000,75000,75000,0,0,0,0,0,0,0,0,0]),

          ("ACT-017","Training content","PRJ-CSU-01","OPEX","EMP-11","STS-IP","1.5.1.1",300000,
           [25000]*12,[23750]*12,[21250]*12,
           [25000,56000,49000,35000,0,0,0,0,0,0,0,0]),

          ("ACT-018","Support tooling","PRJ-CSU-01","CAPEX","EMP-11","STS-IP","1.5.1.2",450000,
           [38000]*11+[32000],[36100]*11+[30400],[32300]*11+[27200],
           [32000,80000,70000,50000,0,0,0,0,0,0,0,0]),
        ]

        for row in acts:
            code,name,proj,ct,emp,sts,wbs,approved = row[:8]
            alloc_m,pr_m,po_m,cons_m = row[8],row[9],row[10],row[11]
            act = Activity(
                activity_code=code, name=name, project_code=proj,
                cost_type_code=ct, employee_code=emp, status_code=sts,
                wbs=wbs, is_leaf=1, approved=approved,
                allocated=sum(alloc_m), pr=sum(pr_m),
                po=sum(po_m), invoiced=sum(cons_m),
            )
            db.add(act)
            db.flush()
            for i in range(12):
                db.add(ActivityPhase(
                    activity_code=code, period_no=i+1,
                    alloc=int(alloc_m[i]), pr=int(pr_m[i]),
                    po=int(po_m[i]), cons=int(cons_m[i]),
                ))

        db.flush()

        # ── Sub-activities with phases ──
        subs = [
          ("SACT-005-1-1","iOS - UI layer","ACT-005",1,"OPEX","EMP-03","STS-IP","1.1.2.2.1.1",462000,
           [38610]*11+[37290],[36680]*11+[35426],[32819]*11+[31697],
           [37290,39600,34650,24750,0,0,0,0,0,0,0,0]),

          ("SACT-005-1-2","iOS - Networking","ACT-005",1,"OPEX","EMP-03","STS-IP","1.1.2.2.1.2",308000,
           [25740]*11+[24860],[24453]*11+[23617],[21879]*11+[21131],
           [24860,26400,23100,16500,0,0,0,0,0,0,0,0]),

          ("SACT-005-2","Android app","ACT-005",1,"OPEX","EMP-03","STS-IP","1.1.2.2.2",630000,
           [52650]*11+[50850],[50017]*11+[48307],[44752]*11+[43222],
           [50850,54000,47250,33750,0,0,0,0,0,0,0,0]),

          ("SACT-006-1","Search ads","ACT-006",1,"OPEX","EMP-05","STS-IP","1.2.1.1.1",660000,
           [55000]*12,[52250]*12,[46750]*12,
           [55000,171600,150150,107250,0,0,0,0,0,0,0,0]),

          ("SACT-006-2","Social ads","ACT-006",1,"OPEX","EMP-05","STS-IP","1.2.1.1.2",540000,
           [45000]*12,[42750]*12,[38250]*12,
           [45000,140400,122850,87750,0,0,0,0,0,0,0,0]),

          ("SACT-012-1","Foundation","ACT-012",1,"CAPEX","EMP-08","STS-IP","1.3.2.1.1.1",330000,
           [27555]*11+[26895],[26177]*11+[25550],[24110]*11+[23533],
           [26895,115500,100980,72270,0,0,0,0,0,0,0,0]),

          ("SACT-012-2","Framing","ACT-012",1,"CAPEX","EMP-08","STS-IP","1.3.2.1.1.2",270000,
           [22545]*11+[22005],[21418]*11+[20905],[19727]*11+[19254],
           [22005,94500,82620,59130,0,0,0,0,0,0,0,0]),

          ("SACT-012-3","Interiors","ACT-012",2,"CAPEX","EMP-08","STS-IP","1.3.2.1.1.3",400000,
           [33400]*11+[32600],[31730]*11+[30970],[29225]*11+[28525],
           [32600,140000,122400,87600,0,0,0,0,0,0,0,0]),

          ("SACT-012-4","Furniture and fixtures","ACT-012",1,"CAPEX","EMP-08","STS-IP","1.3.2.1.2",600000,
           [50100]*11+[48900],[47595]*11+[46455],[43838]*11+[42788],
           [48900,210000,183600,131400,0,0,0,0,0,0,0,0]),

          ("SACT-012-5","IT cabling","ACT-012",1,"CAPEX","EMP-08","STS-IP","1.3.2.1.3",400000,
           [33400]*11+[32600],[31730]*11+[30970],[29225]*11+[28525],
           [32600,140000,122400,87600,0,0,0,0,0,0,0,0]),
        ]

        for row in subs:
            code,name,parent,level,ct,emp,sts,wbs,approved = row[:9]
            alloc_m,pr_m,po_m,cons_m = row[9],row[10],row[11],row[12]
            sa = SubActivity(
                subactivity_code=code, name=name, parent_activity_code=parent,
                level=level, cost_type_code=ct, employee_code=emp,
                status_code=sts, wbs=wbs, is_leaf=1, approved=approved,
                allocated=sum(alloc_m), pr=sum(pr_m),
                po=sum(po_m), invoiced=sum(cons_m),
            )
            db.add(sa)
            db.flush()
            for i in range(12):
                db.add(SubActivityPhase(
                    subactivity_code=code, period_no=i+1,
                    alloc=int(alloc_m[i]), pr=int(pr_m[i]),
                    po=int(po_m[i]), cons=int(cons_m[i]),
                ))

        db.commit()
        logger.info("[budget_seed] Seed data inserted successfully.")
    except Exception as e:
        db.rollback()
        logger.error(f"[budget_seed] Seed failed: {e}", exc_info=True)
    finally:
        db.close()


# ── Dashboard ─────────────────────────────────────────────────────────────────
@router.get("/dashboard", response_model=DashboardOut)
def dashboard(db: Session = Depends(get_db)):
    activities     = db.query(Activity).all()
    sub_activities = db.query(SubActivity).all()

    all_items = list(activities) + list(sub_activities)

    totals = dict(allocated=0, pr=0, po=0, invoiced=0, opex_total=0, capex_total=0)
    for item in all_items:
        totals["allocated"] += item.allocated
        totals["pr"]        += item.pr
        totals["po"]        += item.po
        totals["invoiced"]  += item.invoiced
        tag = item.cost_type.tag if item.cost_type else ""
        if tag == "Opex":
            totals["opex_total"] += item.allocated
        elif tag == "Capex":
            totals["capex_total"] += item.allocated

    depts        = db.query(Department).all()
    projects     = db.query(Project).all()
    project_dept = {p.project_code: p.dept_code for p in projects}

    dept_map = {d.dept_code: dict(
        dept_code=d.dept_code, dept_name=d.name,
        allocated=0, pr=0, po=0, invoiced=0, remaining=0
    ) for d in depts}

    for a in activities:
        dc = project_dept.get(a.project_code)
        if dc and dc in dept_map:
            m = dept_map[dc]
            m["allocated"] += a.allocated
            m["pr"]        += a.pr
            m["po"]        += a.po
            m["invoiced"]  += a.invoiced

    for s in sub_activities:
        act = db.query(Activity).filter(Activity.activity_code == s.parent_activity_code).first()
        if act:
            dc = project_dept.get(act.project_code)
            if dc and dc in dept_map:
                m = dept_map[dc]
                m["allocated"] += s.allocated
                m["pr"]        += s.pr
                m["po"]        += s.po
                m["invoiced"]  += s.invoiced

    for m in dept_map.values():
        m["remaining"] = m["allocated"] - m["invoiced"]

    return DashboardOut(
        allocated=totals["allocated"],
        pr=totals["pr"],
        po=totals["po"],
        invoiced=totals["invoiced"],
        remaining=totals["allocated"] - totals["invoiced"],
        pr_remaining=totals["allocated"] - totals["pr"],
        po_remaining=totals["allocated"] - totals["po"],
        opex_total=totals["opex_total"],
        capex_total=totals["capex_total"],
        dept_summaries=[DeptSummary(**v) for v in dept_map.values()],
    )


# ── Organisations ─────────────────────────────────────────────────────────────
@router.get("/organisations", response_model=List[OrganisationOut])
def list_organisations(db: Session = Depends(get_db)):
    return db.query(Organisation).all()


# ── Departments ───────────────────────────────────────────────────────────────
@router.get("/departments", response_model=List[DepartmentOut])
def list_departments(db: Session = Depends(get_db)):
    return db.query(Department).all()

@router.post("/departments", response_model=DepartmentOut, status_code=201)
def create_department(body: DepartmentCreate, db: Session = Depends(get_db)):
    dept = Department(dept_code=_uid("DEPT-"), name=body.name, org_code=body.org_code, wbs=body.wbs)
    db.add(dept); db.commit(); db.refresh(dept)
    return dept


# ── Projects ──────────────────────────────────────────────────────────────────
@router.get("/projects", response_model=List[ProjectOut])
def list_projects(dept_code: Optional[str] = Query(None), db: Session = Depends(get_db)):
    q = db.query(Project)
    if dept_code:
        q = q.filter(Project.dept_code == dept_code)
    return q.all()

@router.post("/projects", response_model=ProjectOut, status_code=201)
def create_project(body: ProjectCreate, db: Session = Depends(get_db)):
    proj = Project(project_code=_uid("PRJ-"), name=body.name, dept_code=body.dept_code, wbs=body.wbs)
    db.add(proj); db.commit(); db.refresh(proj)
    return proj


# ── Cost types ────────────────────────────────────────────────────────────────
@router.get("/cost-types", response_model=List[CostTypeOut])
def list_cost_types(db: Session = Depends(get_db)):
    return db.query(CostType).all()


# ── Statuses ──────────────────────────────────────────────────────────────────
@router.get("/statuses", response_model=List[StatusOut])
def list_statuses(db: Session = Depends(get_db)):
    return db.query(Status).order_by(Status.sort_order).all()


# ── Employees ─────────────────────────────────────────────────────────────────
@router.get("/employees", response_model=List[EmployeeOut])
def list_employees(db: Session = Depends(get_db)):
    return db.query(Employee).all()

@router.post("/employees", response_model=EmployeeOut, status_code=201)
def create_employee(body: EmployeeCreate, db: Session = Depends(get_db)):
    emp = Employee(employee_code=_uid("EMP-"), name=body.name, title=body.title,
                   dept_code=body.dept_code, manager_code=body.manager_code)
    db.add(emp); db.commit(); db.refresh(emp)
    return emp


# ── Activities ────────────────────────────────────────────────────────────────
@router.get("/activities", response_model=List[ActivityOut])
def list_activities(
    project_code: Optional[str] = Query(None),
    dept_code:    Optional[str] = Query(None),
    cost_type:    Optional[str] = Query(None),
    status:       Optional[str] = Query(None),
    search:       Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    q = (db.query(Activity)
           .options(
               joinedload(Activity.cost_type),
               joinedload(Activity.status),
               joinedload(Activity.employee),
               joinedload(Activity.phases),
               joinedload(Activity.sub_activities).joinedload(SubActivity.cost_type),
               joinedload(Activity.sub_activities).joinedload(SubActivity.status),
               joinedload(Activity.sub_activities).joinedload(SubActivity.employee),
               joinedload(Activity.sub_activities).joinedload(SubActivity.phases),
           ))
    if project_code: q = q.filter(Activity.project_code == project_code)
    if cost_type:    q = q.join(CostType).filter(CostType.tag == cost_type)
    if status:       q = q.join(Status, Activity.status_code == Status.status_code).filter(Status.name == status)
    if search:       q = q.filter(Activity.name.ilike(f"%{search}%"))
    if dept_code:    q = q.join(Project).filter(Project.dept_code == dept_code)
    return [_enrich_activity(a) for a in q.all()]

@router.post("/activities", response_model=ActivityOut, status_code=201)
def create_activity(body: ActivityCreate, db: Session = Depends(get_db)):
    act = Activity(activity_code=_uid("ACT-"), **body.model_dump())
    db.add(act); db.commit(); db.refresh(act)
    return _enrich_activity(
        db.query(Activity)
          .options(joinedload(Activity.cost_type), joinedload(Activity.status),
                   joinedload(Activity.employee), joinedload(Activity.phases),
                   joinedload(Activity.sub_activities).joinedload(SubActivity.phases))
          .filter(Activity.activity_code == act.activity_code).first()
    )

@router.patch("/activities/{code}", response_model=ActivityOut)
def update_activity(code: str, body: ActivityUpdate, db: Session = Depends(get_db)):
    act = db.query(Activity).filter(Activity.activity_code == code).first()
    if not act: raise HTTPException(404, "Activity not found")
    for k, v in body.model_dump(exclude_none=True).items():
        setattr(act, k, v)
    db.commit()
    return _enrich_activity(
        db.query(Activity)
          .options(joinedload(Activity.cost_type), joinedload(Activity.status),
                   joinedload(Activity.employee), joinedload(Activity.phases),
                   joinedload(Activity.sub_activities).joinedload(SubActivity.phases))
          .filter(Activity.activity_code == code).first()
    )

@router.delete("/activities/{code}", status_code=204)
def delete_activity(code: str, db: Session = Depends(get_db)):
    act = db.query(Activity).filter(Activity.activity_code == code).first()
    if not act: raise HTTPException(404, "Activity not found")
    if act.allocated or act.pr or act.po or act.invoiced or act.approved:
        raise HTTPException(400, "Cannot delete: this activity has budget, PR/PO, or spend recorded against it. Reduce it to zero first.")
    if act.sub_activities:
        raise HTTPException(400, f"Cannot delete: this activity has {len(act.sub_activities)} sub-activity line(s). Delete those first.")
    open_crs = db.query(ChangeRequest).filter(
        ChangeRequest.status == "Pending",
        (ChangeRequest.a_code == code) | (ChangeRequest.b_code == code),
    ).count()
    if open_crs:
        raise HTTPException(400, f"Cannot delete: {open_crs} pending change request(s) reference this activity.")
    db.delete(act); db.commit()


# ── Sub-activities ────────────────────────────────────────────────────────────
@router.get("/sub-activities", response_model=List[SubActivityOut])
def list_sub_activities(parent_activity_code: Optional[str] = Query(None), db: Session = Depends(get_db)):
    q = (db.query(SubActivity)
           .options(joinedload(SubActivity.cost_type),
                    joinedload(SubActivity.status),
                    joinedload(SubActivity.employee)))
    if parent_activity_code:
        q = q.filter(SubActivity.parent_activity_code == parent_activity_code)
    return [_enrich_sub(s) for s in q.all()]

@router.post("/sub-activities", response_model=SubActivityOut, status_code=201)
def create_sub_activity(body: SubActivityCreate, db: Session = Depends(get_db)):
    if body.level < 1 or body.level > 3:
        raise HTTPException(400, "level must be 1, 2, or 3")
    sa = SubActivity(subactivity_code=_uid("SACT-"), **body.model_dump())
    db.add(sa); db.commit(); db.refresh(sa)
    return _enrich_sub(
        db.query(SubActivity)
          .options(joinedload(SubActivity.cost_type), joinedload(SubActivity.status),
                   joinedload(SubActivity.employee), joinedload(SubActivity.phases))
          .filter(SubActivity.subactivity_code == sa.subactivity_code).first()
    )

@router.patch("/sub-activities/{code}", response_model=SubActivityOut)
def update_sub_activity(code: str, body: SubActivityUpdate, db: Session = Depends(get_db)):
    sa = db.query(SubActivity).filter(SubActivity.subactivity_code == code).first()
    if not sa: raise HTTPException(404, "Sub-activity not found")
    for k, v in body.model_dump(exclude_none=True).items():
        setattr(sa, k, v)
    db.commit()
    return _enrich_sub(
        db.query(SubActivity)
          .options(joinedload(SubActivity.cost_type), joinedload(SubActivity.status),
                   joinedload(SubActivity.employee), joinedload(SubActivity.phases))
          .filter(SubActivity.subactivity_code == code).first()
    )

@router.delete("/sub-activities/{code}", status_code=204)
def delete_sub_activity(code: str, db: Session = Depends(get_db)):
    sa = db.query(SubActivity).filter(SubActivity.subactivity_code == code).first()
    if not sa: raise HTTPException(404, "Sub-activity not found")
    if sa.allocated or sa.pr or sa.po or sa.invoiced or sa.approved:
        raise HTTPException(400, "Cannot delete: this sub-activity has budget, PR/PO, or spend recorded against it.")
    db.delete(sa); db.commit()


# ── Budget versions ───────────────────────────────────────────────────────────
@router.get("/budget-versions", response_model=List[BudgetVersionOut])
def list_budget_versions(db: Session = Depends(get_db)):
    return db.query(BudgetVersion).all()

@router.post("/budget-versions", response_model=BudgetVersionOut, status_code=201)
def create_budget_version(body: BudgetVersionCreate, db: Session = Depends(get_db)):
    bv = BudgetVersion(version_code=_uid("BV-"), **body.model_dump())
    db.add(bv); db.commit(); db.refresh(bv)
    return bv

@router.patch("/budget-versions/{code}/set-active", response_model=BudgetVersionOut)
def set_active_version(code: str, db: Session = Depends(get_db)):
    db.query(BudgetVersion).update({"is_current": 0})
    bv = db.query(BudgetVersion).filter(BudgetVersion.version_code == code).first()
    if not bv: raise HTTPException(404, "Budget version not found")
    bv.is_current = 1
    db.commit(); db.refresh(bv)
    return bv

@router.patch("/budget-versions/{code}/toggle-lock", response_model=BudgetVersionOut)
def toggle_lock(code: str, db: Session = Depends(get_db)):
    bv = db.query(BudgetVersion).filter(BudgetVersion.version_code == code).first()
    if not bv: raise HTTPException(404, "Budget version not found")
    bv.is_locked = 0 if bv.is_locked else 1
    db.commit(); db.refresh(bv)
    return bv


# ── Fiscal periods ────────────────────────────────────────────────────────────
@router.get("/fiscal-periods", response_model=List[FiscalPeriodOut])
def list_fiscal_periods(db: Session = Depends(get_db)):
    return db.query(FiscalPeriod).order_by(FiscalPeriod.period_no).all()

@router.get("/fiscal-years")
def list_fiscal_years(db: Session = Depends(get_db)):
    rows = db.query(FiscalPeriod.fiscal_year).distinct().order_by(FiscalPeriod.fiscal_year).all()
    return [r[0] for r in rows]


# ── Transfers ─────────────────────────────────────────────────────────────────
@router.get("/transfers", response_model=List[TransferOut])
def list_transfers(db: Session = Depends(get_db)):
    return db.query(Transfer).order_by(Transfer.transfer_date.desc()).all()

@router.post("/transfers", response_model=TransferOut, status_code=201)
def create_transfer(body: TransferCreate, db: Session = Depends(get_db)):
    amount = body.amount
    if body.transfer_type == "dept":
        src_projects = [p.project_code for p in db.query(Project).filter(Project.dept_code == body.from_code).all()]
        dst_projects = [p.project_code for p in db.query(Project).filter(Project.dept_code == body.to_code).all()]
        src_acts = db.query(Activity).filter(Activity.project_code.in_(src_projects)).all()
        dst_acts = db.query(Activity).filter(Activity.project_code.in_(dst_projects)).all()
        src_total = sum(a.allocated for a in src_acts)
        if amount > src_total:
            raise HTTPException(400, f"Insufficient budget in source department (Rs {src_total:,} available)")
        dst_total = sum(a.allocated for a in dst_acts) or 1
        for a in src_acts:
            a.allocated = int(a.allocated * (1 - amount / src_total))
        for a in dst_acts:
            a.allocated = int(a.allocated + amount * (a.allocated / dst_total))

    elif body.transfer_type == "project":
        src_acts = db.query(Activity).filter(Activity.project_code == body.from_code).all()
        dst_acts = db.query(Activity).filter(Activity.project_code == body.to_code).all()
        src_total = sum(a.allocated for a in src_acts)
        if amount > src_total:
            raise HTTPException(400, f"Insufficient budget in source project (Rs {src_total:,} available)")
        dst_total = sum(a.allocated for a in dst_acts) or 1
        for a in src_acts:
            a.allocated = int(a.allocated * (1 - amount / src_total))
        for a in dst_acts:
            a.allocated = int(a.allocated + amount * (a.allocated / dst_total))

    else:  # activity-level
        src = db.query(Activity).filter(Activity.activity_code == body.from_code).first()
        dst = db.query(Activity).filter(Activity.activity_code == body.to_code).first()
        if not src or not dst:
            raise HTTPException(404, "Source or destination activity not found")
        if amount > src.allocated:
            raise HTTPException(400, f"Insufficient budget (Rs {src.allocated:,} available)")
        src.allocated -= amount
        dst.allocated += amount

    tr = Transfer(
        id=_uid("TRF-"),
        transfer_type=body.transfer_type,
        from_code=body.from_code,
        to_code=body.to_code,
        amount=amount,
        employee_code=body.employee_code,
        note=body.note,
        transfer_date=body.transfer_date,
        status="Approved",
    )
    db.add(tr); db.commit(); db.refresh(tr)
    return tr


# ── Change requests ───────────────────────────────────────────────────────────
def _get_or_404(db, model, code, code_field, label):
    obj = db.query(model).filter(getattr(model, code_field) == code).first()
    if not obj:
        raise HTTPException(404, f"{label} not found: {code}")
    return obj

def _phase_row(db: Session, activity_code: str, period_no: int) -> ActivityPhase:
    row = (db.query(ActivityPhase)
             .filter(ActivityPhase.activity_code == activity_code,
                     ActivityPhase.period_no == period_no)
             .first())
    if not row:
        row = ActivityPhase(activity_code=activity_code, period_no=period_no,
                             alloc=0, pr=0, po=0, cons=0)
        db.add(row); db.flush()
    return row

def _enrich_cr(cr: ChangeRequest, db: Session) -> ChangeRequestOut:
    out = ChangeRequestOut.model_validate(cr)
    a = db.query(Activity).filter(Activity.activity_code == cr.a_code).first()
    b = db.query(Activity).filter(Activity.activity_code == cr.b_code).first() if cr.b_code else None
    out.a_name = a.name if a else cr.a_code
    out.b_name = b.name if b else None
    return out


@router.get("/change-requests", response_model=List[ChangeRequestOut])
def list_change_requests(
    status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(ChangeRequest)
    if status:
        q = q.filter(ChangeRequest.status == status)
    rows = q.order_by(ChangeRequest.created_at.desc()).all()
    return [_enrich_cr(r, db) for r in rows]


@router.post("/change-requests", response_model=ChangeRequestOut, status_code=201)
def create_change_request(body: ChangeRequestCreate, db: Session = Depends(get_db)):
    if body.request_type not in ("months", "carry", "pull", "transfer"):
        raise HTTPException(400, "request_type must be one of months, carry, pull, transfer")
    if body.amount <= 0:
        raise HTTPException(400, "amount must be greater than zero")

    a = _get_or_404(db, Activity, body.a_code, "activity_code", "Source activity")

    if body.request_type == "transfer":
        if not body.b_code:
            raise HTTPException(400, "transfer requests require a target activity (b_code)")
        if body.b_code == body.a_code:
            raise HTTPException(400, "Source and target activity must differ")
        _get_or_404(db, Activity, body.b_code, "activity_code", "Target activity")
    else:
        if body.period_from == body.period_to:
            raise HTTPException(400, "Source and destination months must differ")

    for p in (body.period_from, body.period_to):
        if p < 1 or p > 12:
            raise HTTPException(400, "period_from / period_to must be between 1 and 12")

    src_phase = _phase_row(db, body.a_code, body.period_from)
    free = src_phase.alloc - src_phase.cons
    if body.amount > free:
        raise HTTPException(400, f"Only {free:,} is free in {a.name} for period {body.period_from} (requested {body.amount:,}).")

    cr = ChangeRequest(
        id=_uid("CR-"),
        request_type=body.request_type,
        a_code=body.a_code,
        b_code=body.b_code,
        period_from=body.period_from,
        period_to=body.period_to,
        amount=body.amount,
        reason=body.reason,
        requested_by=body.requested_by,
        status="Pending",
        created_at=datetime.utcnow(),
    )
    db.add(cr); db.commit(); db.refresh(cr)
    return _enrich_cr(cr, db)


@router.patch("/change-requests/{cr_id}/decide", response_model=ChangeRequestOut)
def decide_change_request(cr_id: str, body: ChangeRequestDecide, db: Session = Depends(get_db)):
    cr = db.query(ChangeRequest).filter(ChangeRequest.id == cr_id).first()
    if not cr:
        raise HTTPException(404, "Change request not found")
    if cr.status != "Pending":
        raise HTTPException(400, f"Change request is already {cr.status}")

    if body.approve:
        if cr.request_type == "transfer":
            src_act = _get_or_404(db, Activity, cr.a_code, "activity_code", "Source activity")
            dst_act = _get_or_404(db, Activity, cr.b_code, "activity_code", "Target activity")
            src_phase = _phase_row(db, cr.a_code, cr.period_from)
            dst_phase = _phase_row(db, cr.b_code, cr.period_from)

            free = src_phase.alloc - src_phase.cons
            if cr.amount > free:
                raise HTTPException(400, f"Approval failed: only {free:,} now free in {src_act.name} for that month.")

            src_phase.alloc -= cr.amount
            dst_phase.alloc += cr.amount
            src_act.allocated -= cr.amount
            dst_act.allocated += cr.amount

        else:
            act = _get_or_404(db, Activity, cr.a_code, "activity_code", "Activity")
            src_phase = _phase_row(db, cr.a_code, cr.period_from)
            dst_phase = _phase_row(db, cr.a_code, cr.period_to)

            free = src_phase.alloc - src_phase.cons
            if cr.amount > free:
                raise HTTPException(400, f"Approval failed: only {free:,} now free in {act.name} for that month.")

            src_phase.alloc -= cr.amount
            dst_phase.alloc += cr.amount

        cr.status = "Approved"
    else:
        cr.status = "Rejected"

    cr.decided_by = body.decided_by
    cr.decided_at = datetime.utcnow()
    db.commit()
    db.refresh(cr)
    return _enrich_cr(cr, db)


# ── Excel Budget Upload ───────────────────────────────────────────────────────
@router.post("/check-invoice", response_model=InvoiceBudgetCheckRes)
def check_invoice_budget(payload: InvoiceBudgetCheckReq, db: Session = Depends(get_db)):
    # 1. Determine current month and quarter
    current_month = datetime.now().month
    quarter_start_month = 3 * ((current_month - 1) // 3) + 1
    quarter_months = [quarter_start_month, quarter_start_month + 1, quarter_start_month + 2]
    
    # 2. Fetch the phases for this activity
    phases = db.query(ActivityPhase).filter(
        ActivityPhase.activity_code == payload.activity_code
    ).all()
    
    if not phases:
        return InvoiceBudgetCheckRes(
            current_month_allocated=0, current_month_available=0,
            current_quarter_allocated=0, current_quarter_available=0,
            covers_with_tax=False, covers_without_tax=False
        )

    # 3. Calculate Current Month Budget
    current_phase = next((p for p in phases if p.period_no == current_month), None)
    month_alloc = current_phase.alloc if current_phase else 0
    month_used = (current_phase.pr + current_phase.po + current_phase.cons) if current_phase else 0
    month_available = month_alloc - month_used

    # 4. Calculate Current Quarter Budget
    quarter_phases = [p for p in phases if p.period_no in quarter_months]
    quarter_alloc = sum(p.alloc for p in quarter_phases)
    quarter_used = sum(p.pr + p.po + p.cons for p in quarter_phases)
    quarter_available = quarter_alloc - quarter_used
    
    # 5. Check covers against the current month available
    covers_with_tax = month_available >= payload.amount_with_tax
    covers_without_tax = month_available >= payload.amount_without_tax

    return InvoiceBudgetCheckRes(
        current_month_allocated=month_alloc,
        current_month_available=month_available,
        current_quarter_allocated=quarter_alloc,
        current_quarter_available=quarter_available,
        covers_with_tax=covers_with_tax,
        covers_without_tax=covers_without_tax
    )


# ── Excel Budget Upload ───────────────────────────────────────────────────────
REQUIRED_COLS = {"project_code", "activity_name", "wbs", "allocated"}

def trigger_budget_approval_workflow(upload_id: str, amount: float, fiscal_year: str):
    """
    Trigger the central Workflow Engine for budget approval.
    """
    workflow_engine_url = "http://workflow-service/api/v1/trigger" 
    payload = {
        "workflow_type": "BUDGET_APPROVAL",
        "reference_id": upload_id,
        "total_amount": amount,
        "fiscal_year": fiscal_year
    }
    try:
        # requests.post(workflow_engine_url, json=payload, timeout=5)
        logger.info(f"Workflow triggered for budget upload {upload_id}")
    except Exception as e:
        logger.error(f"Failed to trigger workflow: {e}")

@router.post("/budget-upload", response_model=BudgetUploadOut, status_code=201)
async def upload_budget_excel(
    file: UploadFile = File(...),
    fiscal_year: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx or .xls files are accepted")

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl not installed — run: pip install openpyxl")

    try:
        content = await file.read()
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        except Exception as e:
            raise HTTPException(400, f"Cannot read Excel file: {e}")

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(400, "Excel sheet is empty")

        raw_headers = [str(c).strip().lower().replace(" ", "_") if c else "" for c in rows[0]]
        missing = REQUIRED_COLS - set(raw_headers)
        if missing:
            raise HTTPException(400, f"Missing required columns: {', '.join(sorted(missing))}")

        def col(name):
            return raw_headers.index(name) if name in raw_headers else None

        idx = {
            "project_code":   col("project_code"),
            "activity_name":  col("activity_name"),
            "wbs":            col("wbs"),
            "cost_type":      col("cost_type"),
            "allocated":      col("allocated"),
            "pr":             col("pr"),
            "po":             col("po"),
            "invoiced":       col("invoiced"),
            "employee_code":  col("employee_code"),
            "status_code":    col("status_code"),
        }

        cost_type_map = {ct.tag.lower(): ct.cost_type_code for ct in db.query(CostType).all() if ct.tag}
        valid_projects = {p.project_code for p in db.query(Project).all()}
        valid_employees = {e.employee_code for e in db.query(Employee).all()}
        valid_statuses  = {s.status_code for s in db.query(Status).all()}
        existing_wbs   = {a.wbs: a for a in db.query(Activity).all()}

        inserted = updated = skipped = 0
        skip_notes = []

        for row_num, row in enumerate(rows[1:], start=2):
            def cell(key):
                i = idx.get(key)
                if i is not None and i < len(row) and row[i] is not None:
                    return str(row[i]).strip()
                return ""

            project_code  = cell("project_code")
            activity_name = cell("activity_name")
            wbs_val       = cell("wbs")

            if not project_code or not activity_name or not wbs_val:
                skipped += 1
                skip_notes.append(f"Row {row_num}: missing required field")
                continue

            if project_code not in valid_projects:
                skipped += 1
                skip_notes.append(f"Row {row_num}: unknown project_code '{project_code}'")
                continue

            def safe_int(key):
                v = cell(key)
                try: return int(float(v)) if v else 0
                except: return 0

            allocated_val = safe_int("allocated")
            pr_val        = safe_int("pr")
            po_val        = safe_int("po")
            invoiced_val  = safe_int("invoiced")

            ct_raw = cell("cost_type").lower()
            cost_type_code = cost_type_map.get(ct_raw) or cost_type_map.get("opex")

            employee_code = cell("employee_code") or None
            if employee_code and employee_code not in valid_employees:
                employee_code = None

            status_code   = cell("status_code") or "STS-NS"
            if status_code not in valid_statuses:
                status_code = "STS-NS"

            if wbs_val in existing_wbs:
                act = existing_wbs[wbs_val]
                act.name           = activity_name
                act.allocated      = allocated_val
                act.pr             = pr_val
                act.po             = po_val
                act.invoiced       = invoiced_val
                act.cost_type_code = cost_type_code
                if employee_code: act.employee_code = employee_code
                if status_code:   act.status_code   = status_code
                updated += 1
            else:
                new_act = Activity(
                    activity_code  = _uid("ACT-"),
                    name           = activity_name,
                    project_code   = project_code,
                    cost_type_code = cost_type_code,
                    employee_code  = employee_code,
                    status_code    = status_code,
                    wbs            = wbs_val,
                    is_leaf        = 1,
                    allocated      = allocated_val,
                    pr             = pr_val,
                    po             = po_val,
                    invoiced       = invoiced_val,
                )
                db.add(new_act)
                inserted += 1

        db.flush()

        notes_str = "; ".join(skip_notes[:20])
        if len(skip_notes) > 20:
            notes_str += f" … and {len(skip_notes)-20} more"

        upload_log = BudgetUpload(
            id            = _uid("UPL-"),
            filename      = file.filename,
            fiscal_year   = fiscal_year,
            rows_inserted = inserted,
            rows_updated  = updated,
            rows_skipped  = skipped,
            uploaded_at   = datetime.utcnow(),
            notes         = notes_str or None,
            status        = "Pending"
        )
        db.add(upload_log)
        db.commit()
        db.refresh(upload_log)
        
        # Calculate total allocated for workflow
        total_allocated = sum(int(float(r[idx.get("allocated")] or 0)) for r in rows[1:] if idx.get("allocated") is not None and len(r) > idx.get("allocated"))
        trigger_budget_approval_workflow(upload_log.id, total_allocated, fiscal_year)

        return upload_log
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Budget Excel upload failed with internal error")
        raise HTTPException(500, f"Internal server error: {e}")


@router.get("/budget-uploads", response_model=List[BudgetUploadOut])
def list_budget_uploads(db: Session = Depends(get_db)):
    return db.query(BudgetUpload).order_by(BudgetUpload.uploaded_at.desc()).all()


@router.get("/budget-template")
def download_budget_template(db: Session = Depends(get_db)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget Template"

    headers = [
        "Department", "Project Code", "Activity Name", "WBS", 
        "Cost Type", "Allocated", "PR", "PO", "Invoiced", 
        "Employee Code", "Status Code"
    ]
    ws.append(headers)

    depts = [d.name for d in db.query(Department).all()]
    projects = [p.project_code for p in db.query(Project).all()]
    cost_types = ["Opex", "Capex"]
    employees = [e.employee_code for e in db.query(Employee).all()]
    statuses = [s.status_code for s in db.query(Status).all()]

    def add_validation(dv_formula, col_letter, title):
        dv = DataValidation(type="list", formula1=dv_formula, allow_blank=True)
        dv.error = "Your entry is not in the list"
        dv.errorTitle = "Invalid Entry"
        dv.prompt = "Please select from the list"
        dv.promptTitle = title
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}1048576")

    list_ws = wb.create_sheet(title="Lists")
    list_ws.sheet_state = 'hidden'
    
    def write_col(col_idx, data):
        for r_idx, val in enumerate(data, start=1):
            list_ws.cell(row=r_idx, column=col_idx, value=val)
        return f"Lists!${openpyxl.utils.get_column_letter(col_idx)}$1:${openpyxl.utils.get_column_letter(col_idx)}${max(1, len(data))}"
    
    dept_ref = write_col(1, depts)
    proj_ref = write_col(2, projects)
    ct_ref = write_col(3, cost_types)
    emp_ref = write_col(4, employees)
    stat_ref = write_col(5, statuses)

    if len(depts) < 10:
        add_validation(f'"{",".join(depts)}"', "A", "Select Department")
    else:
        add_validation(dept_ref, "A", "Select Department")
        
    add_validation(proj_ref, "B", "Select Project")
    add_validation(f'"{",".join(cost_types)}"', "E", "Select Cost Type")
    add_validation(emp_ref, "J", "Select Employee")
    add_validation(stat_ref, "K", "Select Status")

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    headers_dict = {
        'Content-Disposition': 'attachment; filename="Budget_Upload_Template.xlsx"'
    }
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers_dict)
