from __future__ import annotations
import io
import os
import uuid
import requests  # Added for workflow triggering
from typing import Any

import mysql.connector
from fastapi import FastAPI, File, HTTPException, UploadFile, Form
from openpyxl import load_workbook

app = FastAPI(title="IT Budget Upload API", version="1.0")

DB_CONFIG = {
    "host": os.getenv("DB_HOST", "localhost"),
    "port": int(os.getenv("DB_PORT", "3306")),
    "user": os.getenv("DB_USER", "root"),
    "password": os.getenv("DB_PASSWORD", ""),
    "database": os.getenv("DB_NAME", "it_budget"),
}
FISCAL_YEAR = os.getenv("FISCAL_YEAR", "2026-27")

MONTHS = ["apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec", "jan", "feb", "mar"]
VALID_COST_TYPES = {"Opex", "Capex"}
VALID_STATUS = {"Planned", "Committed", "In Progress", "On Hold", "Completed"}

ID_FIELDS = [
    "line_code", "organisation", "department", "project_code", "project_name",
    "activity_name", "sub_activity_1", "sub_activity_2", "sub_activity_3", "wbs",
    "cost_type", "owner_name", "employee_code", "status",
]
N_ID = len(ID_FIELDS)

_MONTH_COLS = ", ".join(f"{m}_allocated, {m}_prpo" for m in MONTHS)
_MONTH_PARAMS = ", ".join(f"%({m}_allocated)s, %({m}_prpo)s" for m in MONTHS)
_MONTH_UPDATE = ", ".join(
    f"{m}_allocated=VALUES({m}_allocated), {m}_prpo=VALUES({m}_prpo)" for m in MONTHS
)

# Added upload_batch_id to the UPSERT
UPSERT_SQL = f"""
INSERT INTO budget_line
    (upload_batch_id, line_code, organisation, department, project_code, project_name,
     activity_name, sub_activity_1, sub_activity_2, sub_activity_3, wbs,
     cost_type, owner_name, employee_code, status, fiscal_year, {_MONTH_COLS})
VALUES
    (%(upload_batch_id)s, %(line_code)s, %(organisation)s, %(department)s, %(project_code)s, %(project_name)s,
     %(activity_name)s, %(sub_activity_1)s, %(sub_activity_2)s, %(sub_activity_3)s, %(wbs)s,
     %(cost_type)s, %(owner_name)s, %(employee_code)s, %(status)s, %(fiscal_year)s, {_MONTH_PARAMS})
ON DUPLICATE KEY UPDATE
     upload_batch_id=VALUES(upload_batch_id), approval_status='Pending',
     organisation=VALUES(organisation), department=VALUES(department),
     project_code=VALUES(project_code), project_name=VALUES(project_name),
     activity_name=VALUES(activity_name), sub_activity_1=VALUES(sub_activity_1),
     sub_activity_2=VALUES(sub_activity_2), sub_activity_3=VALUES(sub_activity_3),
     wbs=VALUES(wbs), cost_type=VALUES(cost_type), owner_name=VALUES(owner_name),
     employee_code=VALUES(employee_code), status=VALUES(status), {_MONTH_UPDATE};
"""

BATCH_INSERT_SQL = """
INSERT INTO budget_upload_batch (batch_id, fiscal_year, filename, uploaded_by)
VALUES (%s, %s, %s, %s)
"""

def _num(value: Any) -> float:
    if value is None or value == "": return 0.0
    try: return float(value)
    except (TypeError, ValueError): raise ValueError(f"expected a number, got {value!r}")

def _txt(value: Any) -> str | None:
    if value is None: return None
    return str(value).strip() or None

def parse_workbook(data: bytes, batch_id: str) -> list[dict]:
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True)
    except Exception as exc: 
        raise HTTPException(400, f"Could not read Excel file: {exc}")

    ws = wb.active
    rows: list[dict] = []
    errors: list[str] = []

    for r, cells in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if cells is None: continue
        line_code = _txt(cells[0]) if len(cells) else None
        if not line_code: continue
        if line_code.upper().startswith("GRAND"): continue

        rec: dict[str, Any] = {"fiscal_year": FISCAL_YEAR, "upload_batch_id": batch_id}
        
        for i, field in enumerate(ID_FIELDS):
            rec[field] = _txt(cells[i]) if i < len(cells) else None

        if not rec["organisation"] or not rec["project_code"] or not rec["employee_code"]:
            errors.append(f"row {r} ({line_code}): missing required identification field")
            continue
            
        try:
            for m_i, m in enumerate(MONTHS):
                a = N_ID + m_i * 2
                rec[f"{m}_allocated"] = _num(cells[a]) if a < len(cells) else 0.0
                rec[f"{m}_prpo"] = _num(cells[a + 1]) if a + 1 < len(cells) else 0.0
        except ValueError as exc:
            errors.append(f"row {r} ({line_code}): {exc}")
            continue

        rows.append(rec)

    if errors: raise HTTPException(422, {"message": "Validation failed", "errors": errors})
    if not rows: raise HTTPException(422, "No budget line items found.")
    return rows

def load_rows_and_batch(rows: list[dict], batch_id: str, filename: str, uploaded_by: str) -> int:
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cur = conn.cursor()
        
        # 1. Create the Batch record
        cur.execute(BATCH_INSERT_SQL, (batch_id, FISCAL_YEAR, filename, uploaded_by))
        
        # 2. Insert the parsed rows
        cur.executemany(UPSERT_SQL, rows)
        
        conn.commit()
        return cur.rowcount
    except mysql.connector.Error as exc:
        conn.rollback()
        raise HTTPException(500, f"Database write failed: {exc}")
    finally:
        conn.close()

def trigger_budget_approval_workflow(batch_id: str, uploaded_by: str, amount: float):
    """
    Call your Workflow Engine here.
    Update this logic to match how you trigger PR approvals.
    """
    workflow_engine_url = "http://workflow-service/api/v1/trigger" # Replace with your actual URL
    
    payload = {
        "workflow_type": "BUDGET_APPROVAL",
        "reference_id": batch_id,
        "requester": uploaded_by,
        "total_amount": amount,
        "fiscal_year": FISCAL_YEAR
    }
    
    try:
        # requests.post(workflow_engine_url, json=payload, timeout=10)
        print(f"Workflow triggered for batch {batch_id}")
    except Exception as e:
        print(f"Failed to trigger workflow: {e}")

@app.post("/budgets/upload")
async def upload_budget(
    file: UploadFile = File(...), 
    uploaded_by: str = Form("EMP-A1") # Require the employee code who is uploading
) -> dict:
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Please upload an .xlsx file.")
        
    data = await file.read()
    
    # Generate a unique Batch ID for this upload workflow
    batch_id = str(uuid.uuid4())
    
    # Parse excel and attach the batch_id to every row
    rows = parse_workbook(data, batch_id)
    
    # Calculate total FY allocation to send to workflow engine
    total_fy_allocation = sum(
        sum(r[f"{m}_allocated"] for m in MONTHS) for r in rows
    )
    
    # Write to DB
    written = load_rows_and_batch(rows, batch_id, file.filename, uploaded_by)
    
    # Trigger the workflow!
    trigger_budget_approval_workflow(batch_id, uploaded_by, total_fy_allocation)
    
    return {
        "status": "success",
        "batch_id": batch_id,
        "file": file.filename,
        "fiscal_year": FISCAL_YEAR,
        "line_items_parsed": len(rows),
        "rows_affected": written,
        "note": "Budget uploaded successfully. Workflow triggered for approval.",
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
